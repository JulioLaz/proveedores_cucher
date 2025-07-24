from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generar_excel(df: pd.DataFrame, sheet_name: str = "Hoja1") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Escribir encabezados y datos
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

    # Bordes y ajuste de ancho
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999")
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
