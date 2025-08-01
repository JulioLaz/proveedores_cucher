import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from datetime import datetime
import time
import warnings
from pathlib import Path
import matplotlib.dates as mdates
from matplotlib.ticker import FuncFormatter
import locale
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import tempfile
import io
import streamlit as st

warnings.filterwarnings('ignore')
plt.style.use('seaborn-v0_8-darkgrid')
sns.set_palette("husl")


class ExcelHybridPremiumGenerator:
    """
    Generador híbrido premium para Streamlit Cloud
    - Gráficos en carpeta temporal
    - Excel final en memoria (BytesIO)
    """

    def __init__(self, excel_all, proveedor_name="PROVEEDOR"):
        """Inicializa el generador híbrido."""
        self.proveedor_name_clean = proveedor_name.replace(" ", "_").upper()
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_filename = f"{self.proveedor_name_clean}_INFORME_PREMIUM_{self.timestamp}.xlsx"

        # Directorio temporal para gráficos
        self.temp_dir = Path(tempfile.mkdtemp())
        self.charts_dir = self.temp_dir / "charts"
        self.charts_dir.mkdir(exist_ok=True)

        st.write(f"🎯 Generador híbrido premium iniciado")
        st.write(f"🖼️  Directorio temporal: {self.charts_dir}")

        self.excel_all = excel_all
        self.configure_premium_style()
        self.load_data()

    def configure_premium_style(self):
        """Configura estilo visual premium para gráficos."""
        try:
            locale.setlocale(locale.LC_ALL, 'es_AR.UTF-8')
        except:
            try:
                locale.setlocale(locale.LC_ALL, 'Spanish_Argentina.1252')
            except:
                locale.setlocale(locale.LC_ALL, '')

        self.colors = {
            'primary': '#1f77b4',
            'secondary': '#ff7f0e', 
            'success': '#2ca02c',
            'warning': '#d62728',
            'info': '#9467bd',
            'accent': '#8c564b',
            'gradient': ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
        }

        plt.rcParams.update({
            'figure.figsize': (18, 10),
            'font.size': 11,
            'font.family': 'Arial',
            'axes.titlesize': 16,
            'axes.labelsize': 12,
            'xtick.labelsize': 11,
            'ytick.labelsize': 11,
            'legend.fontsize': 11,
            'figure.dpi': 300,
            'savefig.dpi': 300,
            'savefig.bbox': 'tight',
            'savefig.facecolor': 'white',
            'axes.grid': True,
            'grid.alpha': 0.3,
            'axes.spines.top': False,
            'axes.spines.right': False
        })

    def load_data(self):
        """Carga todos los datos del Excel desde excel_all en memoria."""
        try:
            # Si excel_all es un ExcelFile o ruta
            self.sheets = {
                'analisis_producto': pd.read_excel(self.excel_all, sheet_name='Análisis por Producto'),
                'top_performers': pd.read_excel(self.excel_all, sheet_name='Top Performers'),
                'analisis_mensual': pd.read_excel(self.excel_all, sheet_name='Análisis Mensual'),
                'mensual_sucursal': pd.read_excel(self.excel_all, sheet_name='Análisis Mensual por Sucursal'),
                'estacional': pd.read_excel(self.excel_all, sheet_name='Análisis Estacional'),
                'estacional_sucursal': pd.read_excel(self.excel_all, sheet_name='Análisis Estacional por Sucursal'),
                'por_sucursal': pd.read_excel(self.excel_all, sheet_name='Análisis por Sucursal')
            }

            self.preprocess_data()

        except Exception as e:
            st.error(f"❌ Error cargando datos: {str(e)}")
            raise

    # -----------------------
    # 🔹 Mantener aquí tus métodos de gráficos y formateo tal como los tienes
    # 🔹 Solo modificaremos la parte final de guardar Excel
    # -----------------------

    def preprocess_data(self):
        """Preprocesa los datos para optimizar visualizaciones."""
        
        # Procesar fechas
        if 'Periodo' in self.sheets['analisis_mensual'].columns:
            self.sheets['analisis_mensual']['Periodo'] = pd.to_datetime(
                self.sheets['analisis_mensual']['Periodo'], format='%Y-%m'
            )
            
        if 'Periodo' in self.sheets['mensual_sucursal'].columns:
            self.sheets['mensual_sucursal']['Periodo'] = pd.to_datetime(
                self.sheets['mensual_sucursal']['Periodo'], format='%Y-%m'
            )
        
        # Orden de meses
        month_order = ['January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        
        if 'Mes' in self.sheets['estacional'].columns:
            self.sheets['estacional']['Mes'] = pd.Categorical(
                self.sheets['estacional']['Mes'], categories=month_order, ordered=True
            )
            
        if 'Mes' in self.sheets['estacional_sucursal'].columns:
            self.sheets['estacional_sucursal']['Mes'] = pd.Categorical(
                self.sheets['estacional_sucursal']['Mes'], categories=month_order, ordered=True
            )
    
    def format_currency_short(self, value):
        """Formatea valores a moneda abreviada."""
        if value >= 1_000_000_000:
            return f"${value/1_000_000_000:.1f}B"
        elif value >= 1_000_000:
            return f"${value/1_000_000:.1f}M"
        elif value >= 1_000:
            return f"${value/1_000:.1f}K"
        else:
            return f"${value:,.0f}"
    
    def format_quantity_short(self, value):
        """Formatea cantidades abreviadas."""
        if value >= 1_000_000:
            return f"{value/1_000_000:.1f}M"
        elif value >= 1_000:
            return f"{value/1_000:.1f}K"
        else:
            return f"{value:,.0f}"
    
    def add_value_labels(self, ax, bars, is_currency=False, rotation=0):
        """Añade etiquetas de valor premium."""
        for bar in bars:
            height = bar.get_height()
            if height != 0:
                if is_currency:
                    label = self.format_currency_short(height)
                else:
                    label = self.format_quantity_short(height)
                
                ax.annotate(label,
                          xy=(bar.get_x() + bar.get_width() / 2, height),
                          xytext=(0, 3),
                          textcoords="offset points",
                          ha='center', va='bottom',
                          fontweight='bold', fontsize=9,
                          rotation=rotation)
    
    def create_chart_02_analisis_producto(self):
        """Crea gráfico premium para análisis por producto."""
            
        df = self.sheets['analisis_producto'].copy()
        df = df.sort_values('Cant. Mes Pico', ascending=True)
        
        fig, ax = plt.subplots(figsize=(14, 10))
        
        bars = ax.barh(range(len(df)), df['Cant. Mes Pico'], 
                      color=self.colors['gradient'][:len(df)],
                      alpha=0.8, edgecolor='white', linewidth=1)
        
        ax.set_yticks(range(len(df)))
        ax.set_yticklabels([desc[:50] + "..." if len(desc) > 50 else desc 
                           for desc in df['Descripción']], fontsize=10)
        
        # Etiquetas con mes pico
        for i, (bar, mes_pico) in enumerate(zip(bars, df['Mes Pico'])):
            width = bar.get_width()
            if width != 0:
                cantidad_label = self.format_quantity_short(width)
                ax.annotate(f"{cantidad_label}\n({mes_pico})",
                          xy=(width, bar.get_y() + bar.get_height() / 2),
                          xytext=(5, 0),
                          textcoords="offset points",
                          ha='left', va='center',
                          fontweight='bold', fontsize=9)
        
        ax.set_title("ANÁLISIS POR PRODUCTO\nMes Pico vs Cantidad Vendida", 
                    fontsize=18, fontweight='bold', pad=20)
        ax.grid(True, alpha=0.3, axis='x')
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, p: self.format_quantity_short(x)))
        
        plt.tight_layout()
        chart_path = self.charts_dir / "02_analisis_producto.png"
        fig.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        
        return str(chart_path)
    
    def create_chart_03_top_performers(self):
        """Crea gráfico premium para top performers."""
        
        print("  📊 Creando gráfico: Top Performers...")
        
        df = self.sheets['top_performers'].copy()
        df_clean = df[df['Ranking'].apply(lambda x: str(x).isdigit())].copy()
        df_clean['Ranking'] = df_clean['Ranking'].astype(int)
        df_clean['Valor/Métrica'] = df_clean['Valor/Métrica'].str.replace(',', '').astype(float)
        df_top = df_clean.head(10)
        
        fig, ax = plt.subplots(figsize=(14, 10))
        
        bars = ax.bar(range(len(df_top)), df_top['Valor/Métrica'], 
                     color=self.colors['gradient'][:len(df_top)],
                     alpha=0.8, edgecolor='white', linewidth=2)
        
        ax.set_xticks(range(len(df_top)))
        ax.set_xticklabels([f"#{rank}\n{desc[:25]}..." if len(desc) > 25 else f"#{rank}\n{desc}" 
                           for rank, desc in zip(df_top['Ranking'], df_top['Descripción'])], 
                          rotation=45, ha='right', fontsize=9)
        
        self.add_value_labels(ax, bars, is_currency=False)
        
        ax.set_title("TOP PERFORMERS\nProductos por Cantidad Vendida Total", 
                    fontsize=18, fontweight='bold', pad=20)
        ax.grid(True, alpha=0.3, axis='y')
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: self.format_quantity_short(x)))
        
        plt.tight_layout()
        chart_path = self.charts_dir / "03_top_performers.png"
        fig.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        
        return str(chart_path)
    
    def create_chart_04_analisis_mensual(self):
        """Crea gráfico premium para análisis mensual."""
        
        print("  📊 Creando gráfico: Análisis Mensual...")
        
        df = self.sheets['analisis_mensual'].copy()
        df = df.sort_values('Periodo')
        
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 12))
        
        # Gráfico 1: Barras cantidad
        bars = ax1.bar(df['Periodo'], df['Cantidad Vendida'], 
                      color=self.colors['primary'], alpha=0.7,
                      edgecolor='white', linewidth=1)
        
        self.add_value_labels(ax1, bars, is_currency=False, rotation=45)
        
        ax1.set_title("CANTIDAD VENDIDA POR PERÍODO", fontsize=14, fontweight='bold')
        ax1.grid(True, alpha=0.3)
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
        ax1.yaxis.set_major_formatter(FuncFormatter(lambda x, p: self.format_quantity_short(x)))
        
        # Gráfico 2: Serie temporal ventas
        ax2.plot(df['Periodo'], df['Ventas Totales'], 
                color=self.colors['secondary'], linewidth=3, marker='o', 
                markersize=6, markerfacecolor='white', 
                markeredgecolor=self.colors['secondary'], markeredgewidth=2)
        
        ax2.fill_between(df['Periodo'], df['Ventas Totales'], 
                        alpha=0.3, color=self.colors['secondary'])
        
        ax2.set_title("EVOLUCIÓN TEMPORAL DE VENTAS", fontsize=14, fontweight='bold')
        ax2.grid(True, alpha=0.3)
        ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
        plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
        ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, p: self.format_currency_short(x)))
        
        # Punto máximo
        max_idx = df['Ventas Totales'].idxmax()
        max_value = df.loc[max_idx, 'Ventas Totales']
        max_date = df.loc[max_idx, 'Periodo']
        
        ax2.annotate(f'Pico: {self.format_currency_short(max_value)}',
                    xy=(max_date, max_value),
                    xytext=(10, 10),
                    textcoords='offset points',
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='yellow', alpha=0.7),
                    arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'))
        
        fig.suptitle("ANÁLISIS MENSUAL COMPLETO\nCantidad y Ventas por Período", 
                    fontsize=18, fontweight='bold', y=0.95)
        
        plt.tight_layout()
        chart_path = self.charts_dir / "04_analisis_mensual.png"
        fig.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        
        return str(chart_path)
    
    def create_chart_05_mensual_sucursal(self):
        """Crea gráfico premium mensual por sucursal."""
        
        print("  📊 Creando gráfico: Mensual por Sucursal...")
        
        df = self.sheets['mensual_sucursal'].copy()
        df = df.sort_values(['Periodo', 'Sucursal'])
        
        pivot_df = df.pivot(index='Periodo', columns='Sucursal', values='Cantidad Total')
        pivot_df = pivot_df.fillna(0)
        
        fig, ax = plt.subplots(figsize=(14, 10))
        
        width = 0.8 / len(pivot_df.columns)
        x = np.arange(len(pivot_df.index))
        
        for i, sucursal in enumerate(pivot_df.columns):
            bars = ax.bar(x + i * width, pivot_df[sucursal], width, 
                         label=sucursal.title(), 
                         color=self.colors['gradient'][i % len(self.colors['gradient'])],
                         alpha=0.8, edgecolor='white', linewidth=1)
            
            # Etiquetas solo para valores significativos
            for bar in bars:
                height = bar.get_height()
                if height > 1000:
                    ax.annotate(self.format_quantity_short(height),
                              xy=(bar.get_x() + bar.get_width() / 2, height),
                              xytext=(0, 3),
                              textcoords="offset points",
                              ha='center', va='bottom',
                              fontweight='bold', fontsize=8, rotation=90)
        
        ax.set_xticks(x + width * (len(pivot_df.columns) - 1) / 2)
        ax.set_xticklabels([date.strftime('%Y-%m') for date in pivot_df.index], 
                          rotation=45, ha='right')
        
        ax.set_title("ANÁLISIS MENSUAL POR SUCURSAL\nCantidad Total por Período y Ubicación", 
                    fontsize=18, fontweight='bold', pad=20)
        ax.grid(True, alpha=0.3, axis='y')
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: self.format_quantity_short(x)))
        
        if len(pivot_df.columns) > 2:
            ax.legend(title='Sucursal', bbox_to_anchor=(1.05, 1), loc='upper left')
        
        plt.tight_layout()
        chart_path = self.charts_dir / "05_mensual_sucursal.png"
        fig.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        
        return str(chart_path)
    
    def create_chart_06_estacional(self):
        """Crea gráfico premium análisis estacional."""
        
        print("  📊 Creando gráfico: Análisis Estacional...")
        
        df = self.sheets['estacional'].copy()
        df = df.sort_values('Mes')
        
        fig, ax = plt.subplots(figsize=(14, 10))
        
        color_map = {
            '🔥 Alta Temporada': self.colors['warning'],
            '➡️ Temporada Media': self.colors['info'],
            '❄️ Baja Temporada': self.colors['primary']
        }
        
        colors = [color_map.get(cat, self.colors['primary']) for cat in df['Categoría Estacional']]
        
        bars = ax.bar(df['Mes'], df['Cantidad Total'], 
                     color=colors, alpha=0.8, 
                     edgecolor='white', linewidth=2)
        
        self.add_value_labels(ax, bars, is_currency=False, rotation=45)
        
        # Indicadores de categoría
        for i, (bar, categoria) in enumerate(zip(bars, df['Categoría Estacional'])):
            height = bar.get_height()
            emoji = categoria.split()[0]
            ax.annotate(emoji,
                      xy=(bar.get_x() + bar.get_width() / 2, height * 0.1),
                      ha='center', va='bottom', fontsize=16)
        
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        ax.set_title("ANÁLISIS ESTACIONAL\nCantidad Total por Mes del Año", 
                    fontsize=18, fontweight='bold', pad=20)
        ax.grid(True, alpha=0.3, axis='y')
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: self.format_quantity_short(x)))
        
        # Línea de tendencia
        x_numeric = range(len(df))
        z = np.polyfit(x_numeric, df['Cantidad Total'], 1)
        p = np.poly1d(z)
        ax.plot(x_numeric, p(x_numeric), "r--", alpha=0.7, linewidth=2)
        
        plt.tight_layout()
        chart_path = self.charts_dir / "06_estacional.png"
        fig.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        
        return str(chart_path)
    
    def create_chart_08_por_sucursal(self):
        """Crea gráfico premium análisis por sucursal."""
        
        print("  📊 Creando gráfico: Por Sucursal...")
        
        df = self.sheets['estacional_sucursal'].copy()
        df = df.sort_values(['Mes', 'Sucursal'])
        
        pivot_df = df.pivot(index='Mes', columns='Sucursal', values='Cantidad Total')
        pivot_df = pivot_df.fillna(0)
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 10))
        
        # Heatmap
        sns.heatmap(pivot_df.T, annot=True, fmt='.0f', cmap='YlOrRd', 
                   ax=ax1, cbar_kws={'label': 'Cantidad Total'},
                   annot_kws={'size': 8})
        ax1.set_title("MAPA DE CALOR\nCantidad por Mes y Sucursal", 
                     fontsize=14, fontweight='bold')
        ax1.set_ylabel('')
        
        # Barras agrupadas
        width = 0.8 / len(pivot_df.columns)
        x = np.arange(len(pivot_df.index))
        
        for i, sucursal in enumerate(pivot_df.columns):
            ax2.bar(x + i * width, pivot_df[sucursal], width, 
                   label=sucursal.title(), 
                   color=self.colors['gradient'][i % len(self.colors['gradient'])],
                   alpha=0.8, edgecolor='white', linewidth=1)
        
        ax2.set_xticks(x + width * (len(pivot_df.columns) - 1) / 2)
        ax2.set_xticklabels(pivot_df.index, rotation=45, ha='right')
        ax2.set_title("COMPARATIVA POR SUCURSAL\nDistribución Mensual", 
                     fontsize=14, fontweight='bold')
        ax2.grid(True, alpha=0.3, axis='y')
        ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, p: self.format_quantity_short(x)))
        ax2.legend(title='Sucursal', bbox_to_anchor=(1.05, 1), loc='upper left')
        
        fig.suptitle("ANÁLISIS ESTACIONAL POR SUCURSAL\nComparativa de Performance Mensual", 
                    fontsize=18, fontweight='bold', y=0.95)
        
        plt.tight_layout()
        chart_path = self.charts_dir / "08_por_sucursal.png"
        fig.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        
        return str(chart_path)
    
    def generate_all_charts(self):
        """Genera todos los gráficos premium."""
        
        print("\n🎨 GENERANDO GRÁFICOS PREMIUM...")
        print("=" * 50)
        
        chart_files = {}
        
        try:
            chart_files['02'] = self.create_chart_02_analisis_producto()
            chart_files['03'] = self.create_chart_03_top_performers()
            chart_files['04'] = self.create_chart_04_analisis_mensual()
            chart_files['05'] = self.create_chart_05_mensual_sucursal()
            chart_files['06'] = self.create_chart_06_estacional()
            chart_files['08'] = self.create_chart_08_por_sucursal()
            
            print(f"✅ {len(chart_files)} gráficos premium generados")
            return chart_files
            
        except Exception as e:
            print(f"❌ Error generando gráficos: {str(e)}")
            return {}

############################################################

    def create_premium_excel_with_images(self, chart_files):
        """Crea Excel premium y lo devuelve como BytesIO."""
        try:
            # Si excel_all es ExcelFile -> leer contenido binario
            if isinstance(self.excel_all, pd.ExcelFile):
                with open(self.excel_all.io, 'rb') as f:
                    excel_bytes = io.BytesIO(f.read())
            elif isinstance(self.excel_all, (str, Path)):
                with open(self.excel_all, 'rb') as f:
                    excel_bytes = io.BytesIO(f.read())
            else:
                raise ValueError("excel_all debe ser ruta o ExcelFile para openpyxl")

            wb = load_workbook(excel_bytes)

            # Aquí puedes llamar a tus funciones: insert_chart_in_sheet, portada, formateo…
            # ...

            # Guardar Excel final en BytesIO
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)

            return output_buffer

        except Exception as e:
            st.error(f"❌ Error creando Excel premium: {str(e)}")
            return None

    def cleanup_temp_files(self):
        """Limpia archivos temporales."""
        import shutil
        try:
            shutil.rmtree(self.temp_dir)
        except:
            pass


def main():
    # Simulación: excel_all podría ser pd.ExcelFile("archivo.xlsx")
    # En Streamlit, normalmente proviene de st.file_uploader
    uploaded_file = st.file_uploader("📂 Subí tu archivo Excel", type=["xlsx"])
    if not uploaded_file:
        return

    excel_all = pd.ExcelFile(uploaded_file)
    generator = ExcelHybridPremiumGenerator(excel_all, proveedor_name="PRODUMEN SA")

    # Generar gráficos y Excel final
    chart_files = {}  # aquí iría generator.generate_all_charts()
    excel_bytes = generator.create_premium_excel_with_images(chart_files)

    if excel_bytes:
        st.success("✅ Informe premium generado con éxito")
        st.download_button(
            label="⬇️ Descargar Informe Premium",
            data=excel_bytes,
            file_name=generator.output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    generator.cleanup_temp_files()


if __name__ == "__main__":
    main()
