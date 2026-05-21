'''
═══════════════════════════════════════════════════════════════════════════════
 crear_excel_ranking.py
═══════════════════════════════════════════════════════════════════════════════
 Genera un archivo Excel (.xlsx) profesional con el ranking de proveedores
 desglosado por familia / subfamilia (salida de
 process_ranking_data_flias_subflias()).

 Características del Excel:
   • Título + período + sello de generación (filas combinadas).
   • Encabezado con color, negrita y texto blanco.
   • Formatos por tipo: $ para importes, % para porcentajes, # para enteros.
   • Bandas alternadas POR PROVEEDOR (separa visualmente cada bloque
     proveedor → familia → subfamilia).
   • Panes congelados (encabezado + columnas clave fijas al hacer scroll).
   • Anchos de columna automáticos.

 Devuelve un objeto BytesIO listo para usar en st.download_button.

 IMPORTANTE (descarga de UN solo click):
   El patrón st.button(...) -> st.download_button(...) SIEMPRE requiere
   2 clicks. Para 1 click hay que generar los bytes (cacheados) y poner
   st.download_button DIRECTO. Ver el snippet al final del archivo.
═══════════════════════════════════════════════════════════════════════════════
 VERSIONADO
───────────────────────────────────────────────────────────────────────────────
 v1.0  (2026-05-19) - Versión inicial. Excel estilizado con openpyxl, bandas
                      por proveedor, formatos $/%/#, autofiltro, panes
                      congelados y anchos automáticos.
 v1.1  (2026-05-19) - Quitado AutoFilter del Excel (a pedido del usuario).
                      Actualizado el orden de columnas y los grupos de
                      formato para incluir:
                        - 'Ranking-proveedor-subfamilia' (entero).
                        - '% Participación Ventas x Familia' (%).
                        - '% Participación Ventas x Proveedor' (%).
═══════════════════════════════════════════════════════════════════════════════
'''

import io
import time
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def crear_excel_ranking_flias_subflias(df, fecha_desde=None, fecha_hasta=None):
    """
    Genera el Excel del ranking de proveedores (familia/subfamilia).

    Parámetros
    ----------
    df : pd.DataFrame
        Ranking ya procesado (salida de process_ranking_data_flias_subflias).
    fecha_desde, fecha_hasta : date | datetime | str | None
        Solo para el título/contexto del reporte (opcionales).

    Retorna
    -------
    io.BytesIO | None
        Buffer del .xlsx listo para st.download_button. None si hay error.
    """
    print(f"\n{'='*70}")
    print(f"📊 GENERANDO EXCEL — RANKING PROVEEDORES (FLIA/SUBFLIA)")
    print(f"{'='*70}")
    inicio = time.time()

    try:
        if df is None or len(df) == 0:
            print(f"   ⚠️  DataFrame vacío. No se genera Excel.")
            return None

        df = df.copy()

        # === ORDEN DE COLUMNAS PARA EL REPORTE (defensivo, v1.1) ===
        orden_deseado = [
            'Ranking', 'Ranking-proveedor-subfamilia',
            'Proveedor', 'ID Proveedor', 'Familia', 'Subfamilia',
            '% Participación Ventas',
            '% Participación Ventas x Familia',
            '% Participación Ventas x Proveedor',
            'Venta Total', 'Costo Total', 'Utilidad',
            'Rentabilidad %', '% Participación Utilidad',
            'Presupuesto', '% Participación Presupuesto',
            'Cantidad Vendida', 'Artículos',
            'Art. con Exceso', 'Costo Exceso', 'Art. Sin Stock'
        ]
        cols = [c for c in orden_deseado if c in df.columns] + \
               [c for c in df.columns if c not in orden_deseado]
        df = df[cols]
        print(f"   📋 Columnas: {len(cols)}  |  Filas: {len(df):,}")

        # === GRUPOS DE FORMATO POR NOMBRE DE COLUMNA (v1.1) ===
        cols_dinero = {'Venta Total', 'Costo Total', 'Utilidad',
                       'Presupuesto', 'Costo Exceso'}
        cols_pct = {'Rentabilidad %', '% Participación Ventas',
                    '% Participación Ventas x Familia',
                    '% Participación Ventas x Proveedor',
                    '% Participación Utilidad',
                    '% Participación Presupuesto'}
        cols_entero = {'Ranking', 'Ranking-proveedor-subfamilia',
                       'ID Proveedor', 'Cantidad Vendida', 'Artículos',
                       'Art. con Exceso', 'Art. Sin Stock'}

        FMT_DINERO = '"$"#,##0'
        FMT_PCT = '0.000"%"'
        FMT_ENTERO = '#,##0'

        # === ESCRIBIR DATA EN BUFFER (header en fila 3) ===
        output = io.BytesIO()
        START_ROW = 2  # 0-indexed -> encabezado en fila Excel 3

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Ranking',
                        startrow=START_ROW)
            ws = writer.sheets['Ranking']

            n_cols = len(df.columns)
            n_rows = len(df)
            ultima_col = get_column_letter(n_cols)
            fila_header = START_ROW + 1            # Excel 3
            fila_data_ini = fila_header + 1        # Excel 4
            fila_data_fin = fila_header + n_rows

            # --- Paleta ---
            azul_header = PatternFill('solid', fgColor='1F3B57')
            banda_a = PatternFill('solid', fgColor='FFFFFF')
            banda_b = PatternFill('solid', fgColor='EAF1F8')
            fuente_header = Font(bold=True, color='FFFFFF', size=11)
            fuente_titulo = Font(bold=True, color='1F3B57', size=15)
            fuente_sub = Font(italic=True, color='5A6B7B', size=10)
            borde_fino = Border(*[Side(style='thin', color='D0D7DE')] * 4)
            centro = Alignment(horizontal='center', vertical='center')
            izq = Alignment(horizontal='left', vertical='center')

            # --- Título y subtítulo (filas 1 y 2) ---
            ws.merge_cells(f'A1:{ultima_col}1')
            ws['A1'] = 'RANKING DE PROVEEDORES — DESGLOSE FAMILIA / SUBFAMILIA'
            ws['A1'].font = fuente_titulo
            ws['A1'].alignment = centro

            def _fmt_fecha(f):
                if f is None:
                    return None
                if isinstance(f, str):
                    return f
                try:
                    return f.strftime('%d/%m/%Y')
                except Exception:
                    return str(f)

            fd, fh = _fmt_fecha(fecha_desde), _fmt_fecha(fecha_hasta)
            periodo = f'Período: {fd} a {fh}  |  ' if fd and fh else ''
            sello = datetime.now().strftime('%d/%m/%Y %H:%M')
            ws.merge_cells(f'A2:{ultima_col}2')
            ws['A2'] = f'{periodo}Generado: {sello}'
            ws['A2'].font = fuente_sub
            ws['A2'].alignment = centro

            # --- Encabezado (fila 3) ---
            for c_idx in range(1, n_cols + 1):
                cell = ws.cell(row=fila_header, column=c_idx)
                cell.fill = azul_header
                cell.font = fuente_header
                cell.alignment = centro
                cell.border = borde_fino

            # --- Filas de datos: bandas por proveedor + formatos + bordes ---
            col_proveedor = None
            if 'Proveedor' in df.columns:
                col_proveedor = list(df.columns).index('Proveedor')

            proveedor_actual = None
            usar_a = True
            for r in range(n_rows):
                fila_xl = fila_data_ini + r

                if col_proveedor is not None:
                    prov = df.iat[r, col_proveedor]
                    if prov != proveedor_actual:
                        proveedor_actual = prov
                        usar_a = not usar_a
                fill = banda_a if usar_a else banda_b

                for c_idx, col_name in enumerate(df.columns, start=1):
                    cell = ws.cell(row=fila_xl, column=c_idx)
                    cell.fill = fill
                    cell.border = borde_fino

                    if col_name in cols_dinero:
                        cell.number_format = FMT_DINERO
                        cell.alignment = Alignment(horizontal='right',
                                                   vertical='center')
                    elif col_name in cols_pct:
                        cell.number_format = FMT_PCT
                        cell.alignment = Alignment(horizontal='right',
                                                   vertical='center')
                    elif col_name in cols_entero:
                        cell.number_format = FMT_ENTERO
                        cell.alignment = centro
                    else:
                        cell.alignment = izq

            # --- AutoFiltro: QUITADO en v1.1 ---

            # --- Panes congelados (header + columnas clave fijas hasta Subfamilia) ---
            col_freeze = 7 if n_cols >= 7 else 1  # fija hasta 'Subfamilia' (col 6) inclusive
            ws.freeze_panes = ws.cell(row=fila_data_ini, column=col_freeze)

            # --- Anchos automáticos ---
            for c_idx, col_name in enumerate(df.columns, start=1):
                serie = df.iloc[:, c_idx - 1].astype(str)
                max_dato = serie.map(len).max() if len(serie) else 0
                ancho = max(len(str(col_name)), int(max_dato)) + 4
                ws.column_dimensions[get_column_letter(c_idx)].width = \
                    min(max(ancho, 12), 45)

            # --- Alturas ---
            ws.row_dimensions[1].height = 26
            ws.row_dimensions[2].height = 18
            ws.row_dimensions[fila_header].height = 22

        output.seek(0)

        tiempo = time.time() - inicio
        n_prov = df['Proveedor'].nunique() if 'Proveedor' in df.columns else 0
        print(f"   🏢 Proveedores: {n_prov}  |  Filas: {len(df):,}")
        print(f"   ⏱️  Tiempo: {tiempo:.2f}s")
        print(f"   ✅ EXCEL GENERADO EXITOSAMENTE")
        print(f"{'='*70}\n")
        return output

    except Exception as e:
        print(f"   ❌ ERROR generando Excel: {type(e).__name__}: {e}")
        print(f"{'='*70}\n")
        return None