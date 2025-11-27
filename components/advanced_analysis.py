"""
Componente: Análisis Avanzado
Análisis por familia, subfamilia, sucursal y clasificación ABC
Autor: Julio A. Lazarte
"""

import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ============================================
# CONFIGURACIÓN DE MÉTRICAS
# ============================================

METRICAS_OPCIONES = {
    "Ventas": "precio_total",
    "Utilidad": "utilidad",
    "Margen %": "margen_porcentual",
    "Cantidad": "cantidad_total",
    "Participación %": "participacion"
}


# ============================================
# FUNCIONES DE PROCESAMIENTO
# ============================================

def prepare_category_stats(df, group_col):
    """
    Prepara estadísticas agregadas por categoría (familia/subfamilia/sucursal)
    
    Args:
        df (DataFrame): DataFrame con datos
        group_col (str): Columna por la cual agrupar
    
    Returns:
        DataFrame: Estadísticas agregadas
    """
    stats = df.groupby(group_col).agg({
        'precio_total': 'sum',
        'utilidad': 'sum',
        'margen_porcentual': 'mean',
        'cantidad_total': 'sum'
    }).round(2)
    
    stats['participacion'] = (stats['precio_total'] / stats['precio_total'].sum() * 100).round(1)
    
    if group_col == 'sucursal':
        stats['tickets'] = df.groupby(group_col).size()
    
    return stats


def get_format_and_label(columna, stats_df):
    """
    Obtiene formato y etiquetas según métrica
    
    Args:
        columna (str): Nombre de la columna
        stats_df (DataFrame): DataFrame con estadísticas
    
    Returns:
        tuple: (formato, texto_etiqueta)
    """
    if columna in ['precio_total', 'utilidad']:
        formato = "${:,.0f}"
    elif 'margen' in columna or 'participa' in columna:
        formato = "{:,.1f}%"
    else:
        formato = "{:,.0f}"
    
    texto_etiqueta = stats_df[columna].map(formato.format)
    return formato, texto_etiqueta


# ============================================
# GRÁFICOS DE DISTRIBUCIÓN
# ============================================

def render_pie_chart(stats_df, columna, metrica_nombre, titulo_emoji, categoria):
    """
    Renderiza gráfico de pastel con distribución por categoría
    
    Args:
        stats_df (DataFrame): Estadísticas agregadas
        columna (str): Columna a graficar
        metrica_nombre (str): Nombre legible de la métrica
        titulo_emoji (str): Emoji para el título
        categoria (str): Nombre de la categoría
    """
    # Calcular participación para pulls dinámicos
    participacion = (stats_df[columna] / stats_df[columna].sum()) * 100
    pulls = participacion.apply(lambda x: 0.12 if x < 5 else 0.04 if x < 15 else 0.01).tolist()
    
    # Modo de texto según tipo de métrica
    text_mode = 'percent+label' if 'porcentual' in columna or 'participa' in columna else 'label+value'
    
    fig = px.pie(
        stats_df,
        values=columna,
        names=stats_df.index,
        title=f"{titulo_emoji} Distribución de {metrica_nombre} por {categoria}",
        hole=0.35
    )
    
    fig.update_traces(
        textinfo=text_mode,
        textposition='inside',
        pull=pulls,
        marker=dict(line=dict(width=0)),
        hovertemplate=f"<b>%{{label}}</b><br>{metrica_nombre}: %{{value:,.0f}}<br>Participación: %{{percent}}<extra></extra>"
    )
    
    fig.update_layout(
        title_font=dict(size=18, color='#454448', family='Arial Black'),
        title_x=0.08,
        legend=dict(
            bgcolor='rgba(0,0,0,0)',
            bordercolor='rgba(0,0,0,0)',
            font=dict(size=11)
        ),
        showlegend=True,
        margin=dict(t=60, b=30, l=10, r=10)
    )
    
    st.plotly_chart(fig, width="stretch")


def render_bar_chart(stats_df, columna, texto_etiqueta, metrica_nombre, categoria):
    """
    Renderiza gráfico de barras por categoría
    
    Args:
        stats_df (DataFrame): Estadísticas agregadas
        columna (str): Columna a graficar
        texto_etiqueta (Series): Etiquetas formateadas
        metrica_nombre (str): Nombre legible de la métrica
        categoria (str): Nombre de la categoría
    """
    df_bar = stats_df.reset_index()
    
    fig = px.bar(
        df_bar,
        x=categoria,
        y=columna,
        color=columna,
        text=texto_etiqueta,
        title=f"📊 {metrica_nombre} por {categoria.title()}",
        color_continuous_scale='Viridis'
    )
    
    fig.update_traces(
        textposition='outside',
        hovertemplate=f"<b>%{{x}}</b><br>{metrica_nombre}: %{{text}}<extra></extra>"
    )
    
    fig.update_layout(
        title_font=dict(size=18, color='#454448', family='Arial Black'),
        title_x=0.08,
        xaxis_title=None,
        yaxis_title=None,
        coloraxis_showscale=False,
        margin=dict(t=70, b=40, l=30, r=20)
    )
    
    fig.update_yaxes(showticklabels=False)
    st.plotly_chart(fig, width="stretch")


# ============================================
# SECCIONES DE ANÁLISIS
# ============================================

def render_category_analysis(df, group_col, titulo, emoji, key_suffix=""):
    """
    Renderiza análisis completo por categoría (familia/subfamilia/sucursal)
    
    Args:
        df (DataFrame): DataFrame con datos
        group_col (str): Columna por la cual agrupar
        titulo (str): Título de la sección
        emoji (str): Emoji del título
        key_suffix (str): Sufijo para keys únicos de widgets
    """
    if group_col not in df.columns or not df[group_col].notna().any():
        return
    
    # Preparar datos
    stats = prepare_category_stats(df, group_col)
    
    # Header con selector
    col1, col2 = st.columns([3, 2])
    with col1:
        st.markdown(f"### {emoji} {titulo}")
    with col2:
        metrica_seleccionada = st.selectbox(
            "Selecciona una métrica:",
            list(METRICAS_OPCIONES.keys()),
            index=0,
            key=f"metrica_{key_suffix}"
        )
    
    # Obtener columna y formato
    columna = METRICAS_OPCIONES[metrica_seleccionada]
    stats = stats.sort_values(columna, ascending=False)
    formato, texto_etiqueta = get_format_and_label(columna, stats)
    
    # Gráficos
    col1, col2 = st.columns(2)
    
    with col1:
        render_pie_chart(stats, columna, metrica_seleccionada, "🥧", titulo.split()[-1])
    
    with col2:
        render_bar_chart(stats, columna, texto_etiqueta, metrica_seleccionada, group_col)


# ============================================
# ANÁLISIS ABC
# ============================================

def prepare_abc_data(df):
    """
    Prepara datos para análisis ABC
    
    Args:
        df (DataFrame): DataFrame con datos
    
    Returns:
        tuple: (productos_abc, abc_counts, abc_ventas)
    """
    productos_abc = df.groupby(['idarticulo', 'descripcion']).agg({
        'precio_total': 'sum',
        'utilidad': 'sum'
    }).sort_values('precio_total', ascending=False)
    
    productos_abc['participacion_acum'] = (
        productos_abc['precio_total'].cumsum() /
        productos_abc['precio_total'].sum() * 100
    )
    
    # Clasificación ABC
    def categorizar_abc(part):
        if part <= 80:
            return 'A (Alto valor)'
        elif part <= 95:
            return 'B (Valor medio)'
        else:
            return 'C (Bajo valor)'
    
    productos_abc['categoria_abc'] = productos_abc['participacion_acum'].apply(categorizar_abc)
    
    # Datos agregados
    abc_counts = productos_abc['categoria_abc'].value_counts().sort_index()
    abc_ventas = productos_abc.groupby('categoria_abc')['precio_total'].sum().sort_index()
    
    return productos_abc, abc_counts, abc_ventas


def render_abc_charts(abc_counts, abc_ventas):
    """
    Renderiza gráficos de análisis ABC
    
    Args:
        abc_counts (Series): Conteo de productos por categoría
        abc_ventas (Series): Ventas por categoría
    """
    col1, col2 = st.columns(2)
    
    with col1:
        fig = px.bar(
            x=abc_counts.index,
            y=abc_counts.values,
            color=abc_counts.values,
            text=abc_counts.values,
            title="📦 Cantidad de Productos por Categoría ABC",
            labels={'x': 'Categoría ABC', 'y': 'Cantidad'},
            color_continuous_scale='Blues'
        )
        fig.update_traces(
            textposition='outside',
            hovertemplate="<b>%{x}</b><br>Cantidad: %{y}<extra></extra>"
        )
        fig.update_layout(
            title_font=dict(size=18, color='#2c2c2c', family='Arial Black'),
            title_x=0.08,
            xaxis_title=None,
            yaxis_title=None,
            coloraxis_showscale=False,
            height=400,
            margin=dict(t=60, b=40, l=30, r=20)
        )
        st.plotly_chart(fig, width="stretch")
    
    with col2:
        fig = px.pie(
            values=abc_ventas.values,
            names=abc_ventas.index,
            title="💰 Participación de Ventas por Categoría ABC",
            hole=0.35
        )
        fig.update_traces(
            textinfo='percent+label',
            textposition='inside',
            marker=dict(line=dict(width=0)),
            hovertemplate="<b>%{label}</b><br>Ventas: %{value:$,.0f}<br>Participación: %{percent}<extra></extra>"
        )
        fig.update_layout(
            title_font=dict(size=18, color='#2c2c2c', family='Arial Black'),
            title_x=0.08,
            legend=dict(
                bgcolor='rgba(0,0,0,0)',
                bordercolor='rgba(0,0,0,0)',
                font=dict(size=11)
            ),
            height=400,
            margin=dict(t=60, b=30, l=10, r=10)
        )
        st.plotly_chart(fig, width="stretch")


def inject_insights_css():
    """Inyecta estilos CSS para insights"""
    st.markdown("""
    <style>
        .insight-box, .warning-box, .success-box {
            border-radius: 12px;
            padding: 1.2rem;
            margin: 1rem 0;
            font-size: 0.95rem;
            line-height: 1.6;
            background-color: #ffffff;
            box-shadow: 0 4px 12px rgba(0,0,0,0.05);
            border-left: 5px solid #2a5298;
        }
        .warning-box {
            border-left-color: #ffc107;
            background-color: #fff9e6;
        }
        .success-box {
            border-left-color: #28a745;
            background-color: #e9f7ef;
        }
        .insight-box {
            border-left-color: #17a2b8;
            background-color: #eef9fc;
        }
    </style>
    """, unsafe_allow_html=True)


# ============================================
# RECOMENDACIONES ESTRATÉGICAS
# ============================================

def generate_strategic_recommendations(productos_abc, abc_ventas, metrics):
    """
    Genera recomendaciones estratégicas basadas en ABC y métricas
    
    Args:
        productos_abc (DataFrame): DataFrame con clasificación ABC
        abc_ventas (Series): Ventas por categoría ABC
        metrics (dict): Métricas generales
    
    Returns:
        dict: Recomendaciones por prioridad
    """
    recomendaciones = {
        'criticas': [],
        'medias': [],
        'bajas': []
    }
    
    # Productos A
    productos_a = productos_abc[productos_abc['categoria_abc'] == 'A (Alto valor)']
    if not productos_a.empty:
        ventas_a = abc_ventas.get('A (Alto valor)', 0)
        porcentaje_a = ventas_a / abc_ventas.sum() * 100
        recomendaciones['criticas'].append(
            f"🔺 **Productos A:** {len(productos_a)} productos generan el {porcentaje_a:.1f}% de las ventas. Priorizá disponibilidad y promoción."
        )
    
    # Margen bajo
    if metrics['margen_promedio'] < 20:
        recomendaciones['criticas'].append(
            f"🔴 **Margen bajo ({metrics['margen_promedio']:.1f}%):** Revisar precios y negociar con proveedores."
        )
    elif metrics['margen_promedio'] >= 30:
        recomendaciones['bajas'].append(
            f"✅ **Margen saludable:** Excelente rentabilidad promedio ({metrics['margen_promedio']:.1f}%). ¡Seguir así!"
        )
    
    # Diversificación
    if metrics['productos_unicos'] < 10:
        recomendaciones['medias'].append(
            f"📈 **Ampliar catálogo:** Solo {metrics['productos_unicos']} productos únicos. Evaluar incorporar nuevas líneas."
        )
    else:
        recomendaciones['bajas'].append(
            f"🟢 **Catálogo variado:** {metrics['productos_unicos']} productos activos. Diversificación saludable."
        )
    
    return recomendaciones


def render_recommendations(recomendaciones):
    """
    Renderiza recomendaciones estratégicas
    
    Args:
        recomendaciones (dict): Recomendaciones por prioridad
    """
    st.markdown("### 💡 Recomendaciones Estratégicas")
    
    if recomendaciones['criticas']:
        st.markdown("#### 🔺 Alta Prioridad")
        for rec in recomendaciones['criticas']:
            st.markdown(f'<div class="insight-box red">{rec}</div>', unsafe_allow_html=True)
    
    if recomendaciones['medias']:
        st.markdown("#### ⚠️ Prioridad Media")
        for rec in recomendaciones['medias']:
            st.markdown(f'<div class="insight-box">{rec}</div>', unsafe_allow_html=True)
    
    if recomendaciones['bajas']:
        st.markdown("#### ✅ Aspectos Positivos")
        for rec in recomendaciones['bajas']:
            st.markdown(f'<div class="insight-box green">{rec}</div>', unsafe_allow_html=True)


# ============================================
# EXPORTACIÓN A EXCEL
# ============================================

def generate_abc_excel(df_abc):
    """
    Genera archivo Excel con clasificación ABC
    
    Args:
        df_abc (DataFrame): DataFrame con clasificación ABC
    
    Returns:
        BytesIO: Archivo Excel en memoria
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Clasificación ABC"
    
    # Escribir datos
    for r_idx, row in enumerate(dataframe_to_rows(df_abc, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Estilos
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    currency_fmt = '"$"#,##0'
    percent_fmt = '0.0"%"'
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999")
    )
    
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        max_length = 0
        for cell in col:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            if isinstance(cell.value, (int, float)):
                if cell.column_letter in ['C', 'D']:
                    cell.number_format = currency_fmt
                elif cell.column_letter == 'E':
                    cell.number_format = percent_fmt
            max_length = max(max_length, len(str(cell.value)))
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================
# FUNCIÓN PRINCIPAL
# ============================================

def show_advanced_analysis(df, metrics, 
                          generar_insight_margen_func=None,
                          generar_insight_cantidad_func=None,
                          generar_insight_ventas_func=None,
                          generar_insight_abc_completo_func=None):
    """
    Muestra análisis avanzado completo
    
    Args:
        df (DataFrame): DataFrame con datos
        metrics (dict): Métricas generales
        generar_insight_margen_func (callable): Función para insights de margen
        generar_insight_cantidad_func (callable): Función para insights de cantidad ABC
        generar_insight_ventas_func (callable): Función para insights de ventas ABC
        generar_insight_abc_completo_func (callable): Función para insights ABC completo
    
    Estructura:
        1. Análisis por familia
        2. Análisis por subfamilia
        3. Análisis por sucursal
        4. Insights de márgenes
        5. Análisis ABC
        6. Recomendaciones estratégicas
        7. Tabla ABC con descarga
    """
    # Validar datos
    if df is None or df.empty:
        st.error("❌ No hay datos disponibles para el análisis avanzado")
        return
    
    # 1. Análisis por Familia
    render_category_analysis(df, 'familia', 
                            'Análisis por Familia de Productos', 
                            '🌿', 
                            'familia')
    
    # 2. Análisis por Subfamilia
    render_category_analysis(df, 'subfamilia', 
                            'Análisis por Subfamilia de Productos', 
                            '🍃', 
                            'subfamilia')
    
    # 3. Análisis por Sucursal
    render_category_analysis(df, 'sucursal', 
                            'Análisis por Sucursal', 
                            '🏪', 
                            'sucursal')
    
    # 4. Insights de márgenes (si se proporciona la función)
    if generar_insight_margen_func:
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if 'sucursal' in df.columns:
                df_margenes_suc = df.groupby('sucursal')['margen_porcentual'].mean()
                st.markdown(generar_insight_margen_func(df_margenes_suc, "Sucursal"), 
                          unsafe_allow_html=True)
        
        with col2:
            if 'familia' in df.columns:
                df_margenes_flia = df.groupby('familia')['margen_porcentual'].mean()
                st.markdown(generar_insight_margen_func(df_margenes_flia, "Familia"), 
                          unsafe_allow_html=True)
        
        with col3:
            if 'subfamilia' in df.columns:
                df_margenes_subflia = df.groupby('subfamilia')['margen_porcentual'].mean()
                st.markdown(generar_insight_margen_func(df_margenes_subflia, "Subfamilia"), 
                          unsafe_allow_html=True)
    
    # 5. Análisis ABC
    st.markdown("### 📊 Análisis ABC de Productos")
    productos_abc, abc_counts, abc_ventas = prepare_abc_data(df)
    render_abc_charts(abc_counts, abc_ventas)
    
    # Inyectar CSS para insights
    inject_insights_css()
    
    # Insights ABC (si se proporcionan las funciones)
    col1, col2 = st.columns(2)
    
    with col1:
        if generar_insight_cantidad_func:
            st.markdown(generar_insight_cantidad_func(abc_counts), unsafe_allow_html=True)
    
    with col2:
        if generar_insight_ventas_func:
            st.markdown(generar_insight_ventas_func(abc_ventas), unsafe_allow_html=True)
    
    if generar_insight_abc_completo_func:
        st.markdown(generar_insight_abc_completo_func(abc_counts, abc_ventas), 
                   unsafe_allow_html=True)
    
    # 6. Recomendaciones estratégicas
    recomendaciones = generate_strategic_recommendations(productos_abc, abc_ventas, metrics)
    render_recommendations(recomendaciones)
    
    # 7. Tabla ABC con descarga
    st.markdown("### 📋 Detalle de Clasificación ABC")
    
    tabla_abc = productos_abc.reset_index()[[
        'idarticulo', 'descripcion', 'precio_total', 'utilidad', 
        'participacion_acum', 'categoria_abc'
    ]]
    tabla_abc.columns = [
        'ID Artículo', 'Descripción', 'Ventas Totales', 
        'Utilidad', 'Participación Acum. (%)', 'Categoría ABC'
    ]
    
    # Formatear valores
    tabla_abc['Ventas Totales'] = tabla_abc['Ventas Totales'].round(0).astype(int)
    tabla_abc['Utilidad'] = tabla_abc['Utilidad'].round(0).astype(int)
    tabla_abc['Participación Acum. (%)'] = tabla_abc['Participación Acum. (%)'].round(1)
    
    st.dataframe(tabla_abc, width="stretch")
    
    # Botón de descarga
    archivo_excel = generate_abc_excel(tabla_abc)
    st.download_button(
        label="📥 Descargar tabla ABC en Excel",
        data=archivo_excel,
        file_name="clasificacion_abc.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ============================================
# EXPORTAR FUNCIONES PÚBLICAS
# ============================================

__all__ = [
    'show_advanced_analysis',
    'prepare_abc_data',
    'generate_strategic_recommendations'
]