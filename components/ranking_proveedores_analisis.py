# ============================================================================
# MÓDULO: RANKING DE PROVEEDORES - ANÁLISIS COMPLETO
# ============================================================================
"""
Genera análisis completo de ranking de proveedores con:
- 3 Rankings: por Cantidad, Venta, Utilidad
- 3 Detalles por artículo: por Cantidad, Venta, Utilidad
- Formato Excel profesional con estilos
- Corrección de proveedor para artículos de SALTA REFRESCOS
"""

import pandas as pd
import numpy as np
from datetime import datetime
import time
import re
import warnings
from io import BytesIO
import os
warnings.filterwarnings('ignore')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠️ openpyxl no está instalado. Instalar con: pip install openpyxl")

# ============================================================================
# FUNCIONES AUXILIARES DE LOGGING
# ============================================================================

def print_header(text):
    """Imprime un encabezado visual"""
    print("\n" + "="*80)
    print(f"  {text}")
    print("="*80)

def print_subheader(text):
    """Imprime un subencabezado visual"""
    print(f"\n{'─'*80}")
    print(f"  {text}")
    print(f"{'─'*80}")

def print_info(label, value):
    """Imprime información formateada"""
    print(f"  ✓ {label:.<60} {value}")

def print_time(label, seconds):
    """Imprime tiempo de ejecución"""
    print(f"  ⏱  {label:.<60} {seconds:.2f}s")

# ============================================================================
# FUNCIÓN PARA CARGAR SALTA REFRESCOS
# ============================================================================

def cargar_salta_refrescos():
    """
    Carga lista de artículos de Salta Refrescos desde utils/
    
    Returns:
        list: Lista de idarticuloalfa como strings
    """
    try:
        path_salta = os.path.join('utils', 'SALTA_REFRESCOS.csv')
        df_salta = pd.read_csv(path_salta)
        lista_salta = df_salta['idarticuloalfa'].astype(str).str.strip().tolist()
        return lista_salta
    except Exception as e:
        print(f"  ⚠️ Error al cargar SALTA_REFRESCOS.csv: {e}")
        return []

# ============================================================================
# FUNCIONES DE LIMPIEZA DE DATOS
# ============================================================================

def limpiar_caracteres_ilegales(text):
    """
    Limpia caracteres ilegales para Excel (caracteres de control ASCII 0-31 
    excepto tab, newline, carriage return)
    
    Args:
        text: Texto a limpiar
        
    Returns:
        str: Texto limpio
    """
    if pd.isna(text):
        return text
    
    if not isinstance(text, str):
        text = str(text)
    
    # Patrón para caracteres ilegales en Excel
    # Excluye: \t (tab=9), \n (newline=10), \r (carriage return=13)
    illegal_chars = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]')
    
    # Reemplazar caracteres ilegales con espacio
    cleaned_text = illegal_chars.sub(' ', text)
    
    return cleaned_text

def limpiar_dataframe_para_excel(df):
    """
    Limpia todas las columnas de texto de un DataFrame para que sean 
    compatibles con Excel
    
    Args:
        df: DataFrame a limpiar
        
    Returns:
        DataFrame: DataFrame limpio
    """
    df_clean = df.copy()
    
    # Aplicar limpieza a todas las columnas de tipo object (texto)
    for col in df_clean.select_dtypes(include=['object']).columns:
        df_clean[col] = df_clean[col].apply(limpiar_caracteres_ilegales)
    
    return df_clean

# ============================================================================
# FUNCIONES DE FORMATO EXCEL
# ============================================================================

def ajustar_ancho_columnas(ws):
    """
    Ajusta el ancho de las columnas basándose en el contenido
    Deja un espacio adicional antes y después del texto
    
    Args:
        ws: Worksheet de openpyxl
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Ajustar ancho: contenido + espacios adicionales
        adjusted_width = max_length + 4
        ws.column_dimensions[column_letter].width = adjusted_width

def formatear_hoja(ws, nombre_hoja):
    """
    Aplica todo el formateo a una hoja específica:
    - Headers amarillos con altura 30
    - Freeze panes (primera fila + 2 columnas)
    - Formatos numéricos (moneda, número, porcentaje)
    - Autoajuste de columnas
    
    Args:
        ws: Worksheet de openpyxl
        nombre_hoja: Nombre de la hoja (para logging)
    """
    print(f"\n  Formateando hoja: {nombre_hoja}")
    t0 = time.time()
    
    # 1. FORMATO DE ENCABEZADOS (Primera fila)
    # ─────────────────────────────────────────
    # Color amarillo claro
    fill_amarillo = PatternFill(start_color='E6BD5A', end_color='E6C781', fill_type='solid')
    
    # Altura de la primera fila
    ws.row_dimensions[1].height = 30
    
    # Aplicar formato a cada celda del encabezado
    for cell in ws[1]:
        cell.fill = fill_amarillo
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 2. INMOVILIZAR PANELES (Primera fila y 2 primeras columnas)
    # ────────────────────────────────────────────────────────────
    ws.freeze_panes = 'C2'  # Congela hasta columna B (2 columnas) y fila 1
    
    # 3. IDENTIFICAR COLUMNAS Y APLICAR FORMATOS
    # ───────────────────────────────────────────
    columnas_moneda = ['venta_total', 'utilidad_total', 'precio_total', 'costo_total']
    columnas_numero = ['cantidad_total', 'articulos_unicos']
    columnas_porcentaje = ['rentabilidad_pct']
    
    # Obtener índices de columnas
    headers = [cell.value for cell in ws[1]]
    
    # Diccionario para mapear nombre -> índice de columna
    col_indices = {}
    for idx, header in enumerate(headers, start=1):
        if header:
            col_indices[header] = idx
    
    # Aplicar formato moneda (integer)
    for col_name in columnas_moneda:
        if col_name in col_indices:
            col_idx = col_indices[col_name]
            col_letter = get_column_letter(col_idx)
            
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                if cell.value is not None:
                    cell.number_format = '$#,##0'  # Formato moneda sin decimales
    
    # Aplicar formato número (integer)
    for col_name in columnas_numero:
        if col_name in col_indices:
            col_idx = col_indices[col_name]
            col_letter = get_column_letter(col_idx)
            
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                if cell.value is not None:
                    cell.number_format = '#,##0'  # Formato número sin decimales
    
    # Aplicar formato porcentaje
    for col_name in columnas_porcentaje:
        if col_name in col_indices:
            col_idx = col_indices[col_name]
            col_letter = get_column_letter(col_idx)
            
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                if cell.value is not None:
                    cell.number_format = '0.00%'  # Formato porcentaje con 2 decimales
    
    # 4. AJUSTAR ANCHO DE COLUMNAS
    # ─────────────────────────────
    ajustar_ancho_columnas(ws)
    
    t1 = time.time()
    print_time(f"Hoja '{nombre_hoja}' formateada", t1-t0)

# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

def main_ranking_proveedores_analisis(
    df_ventas: pd.DataFrame,
    df_proveedores: pd.DataFrame,
    fecha_desde: str,
    fecha_hasta: str
) -> BytesIO:
    """
    Genera análisis completo de ranking de proveedores con formato Excel
    
    Args:
        df_ventas: DataFrame con ventas del período (completo, sin filtros familia/subfamilia)
        df_proveedores: DataFrame con relación artículo-proveedor
        fecha_desde: Fecha inicio del período (string o date)
        fecha_hasta: Fecha fin del período (string o date)
        
    Returns:
        BytesIO: Archivo Excel en memoria para descarga
    """
    
    print_header("🚀 ANÁLISIS DE RANKING POR PROVEEDORES")
    print(f"  Fecha de ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  {'='*80}")
    
    tiempo_total_inicio = time.time()
    
    # ========================================================================
    # 1. VALIDACIÓN DE DATOS DE ENTRADA
    # ========================================================================
    print_subheader("📊 VALIDACIÓN DE DATOS")
    
    if df_ventas is None or len(df_ventas) == 0:
        print("  ❌ Error: df_ventas está vacío")
        return None
    
    if df_proveedores is None or len(df_proveedores) == 0:
        print("  ❌ Error: df_proveedores está vacío")
        return None
    
    print_info("Registros en df_ventas", f"{len(df_ventas):,}")
    print_info("Registros en df_proveedores", f"{len(df_proveedores):,}")
    print_info("Rango de fechas", f"{fecha_desde} → {fecha_hasta}")
    
    # ========================================================================
    # 2. CORRECCIÓN DE PROVEEDORES - SALTA REFRESCOS
    # ========================================================================
    print_subheader("🔧 CORRECCIÓN DE PROVEEDORES")
    
    print("  Cargando artículos de SALTA REFRESCOS...")
    t0 = time.time()
    
    idartalfa_salta = cargar_salta_refrescos()
    
    if idartalfa_salta:
        print_info("Artículos únicos en SALTA_REFRESCOS", f"{len(idartalfa_salta):,}")
        
        # Preparar df_proveedores
        df_proveedores = df_proveedores.copy()
        
        # Asegurar que idartalfa existe en df_proveedores
        if 'idartalfa' not in df_proveedores.columns:
            print("  ⚠️ Columna 'idartalfa' no encontrada en df_proveedores")
            print("  ⚠️ No se aplicará corrección de SALTA REFRESCOS")
        else:
            df_proveedores['idartalfa'] = df_proveedores['idartalfa'].astype(str).str.strip()
            idartalfa_salta_str = [str(x).strip() for x in idartalfa_salta]
            
            # Actualizar el proveedor
            mask_salta = df_proveedores['idartalfa'].isin(idartalfa_salta_str)
            articulos_actualizados = mask_salta.sum()
            
            df_proveedores.loc[mask_salta, 'proveedor'] = 'SALTA REFRESCOS S.A. (CTES)   '
            
            t1 = time.time()
            print_time("Proveedores actualizados", t1-t0)
            print_info("Artículos actualizados", f"{articulos_actualizados:,}")
    else:
        print("  ⚠️ No se pudo cargar SALTA_REFRESCOS.csv")
        print("  ⚠️ Continuando sin corrección de proveedores")
    
    # ========================================================================
    # 3. MERGE DE DATOS
    # ========================================================================
    print_subheader("🔗 MERGE DE DATOS")
    
    print("  Realizando merge entre df_ventas y df_proveedores...")
    t0 = time.time()
    
    # Merge por idarticulo
    df_merged = df_ventas.merge(
        df_proveedores[['idarticulo', 'idproveedor', 'proveedor']],
        on='idarticulo',
        how='left'
    )
    
    t1 = time.time()
    print_time("Merge completado", t1-t0)
    print_info("Registros después del merge", f"{len(df_merged):,}")
    print_info("Registros sin proveedor", f"{df_merged['proveedor'].isna().sum():,}")
    
    # Filtrar solo registros con proveedor
    df_merged = df_merged[df_merged['proveedor'].notna()].copy()
    print_info("Registros con proveedor asignado", f"{len(df_merged):,}")
    
    # ========================================================================
    # 4. CÁLCULO DE UTILIDAD
    # ========================================================================
    print_subheader("💰 CÁLCULO DE MÉTRICAS")
    
    print("  Calculando utilidad_total...")
    t0 = time.time()
    
    # Verificar columnas necesarias
    if 'venta_total' not in df_merged.columns or 'costo_total' not in df_merged.columns:
        print("  ❌ Error: Columnas 'venta_total' o 'costo_total' no encontradas")
        return None
    
    df_merged['utilidad_total'] = df_merged['venta_total'] - df_merged['costo_total']
    
    t1 = time.time()
    print_time("Utilidad calculada", t1-t0)
    
    # ========================================================================
    # 5. AGRUPACIÓN POR PROVEEDOR
    # ========================================================================
    print_subheader("📊 AGRUPACIÓN Y RANKINGS")
    
    print("  Agrupando datos por proveedor...")
    t0 = time.time()
    
    # Agrupar por proveedor
    df_proveedores_grouped = df_merged.groupby('proveedor').agg({
        'cantidad_vendida': 'sum',
        'venta_total': 'sum',
        'costo_total': 'sum',
        'utilidad_total': 'sum',
        'idarticulo': 'nunique'  # Cantidad de artículos únicos
    }).reset_index()
    
    # Renombrar columnas
    df_proveedores_grouped.columns = ['proveedor', 'cantidad_total', 'venta_total', 
                              'costo_total', 'utilidad_total', 'articulos_unicos']
    
    # Calcular rentabilidad
    df_proveedores_grouped['rentabilidad_pct'] = (
        df_proveedores_grouped['utilidad_total'] / df_proveedores_grouped['costo_total']
    ).round(4)
    
    t1 = time.time()
    print_time("Agrupación completada", t1-t0)
    print_info("Proveedores únicos", f"{len(df_proveedores_grouped):,}")
    
    # ========================================================================
    # 6. CREAR RANKINGS
    # ========================================================================
    print("\n  Creando rankings...")
    t0 = time.time()
    
    # Ranking 1: Por Cantidad Total
    ranking_cantidad = df_proveedores_grouped.copy()
    ranking_cantidad['ranking'] = ranking_cantidad['cantidad_total'].rank(
        ascending=False, method='min'
    ).astype(int)
    ranking_cantidad = ranking_cantidad.sort_values('cantidad_total', ascending=False)
    ranking_cantidad = ranking_cantidad[[
        'ranking', 'proveedor', 'cantidad_total', 'venta_total', 'utilidad_total',
        'rentabilidad_pct', 'articulos_unicos'
    ]]
    
    # Ranking 2: Por Venta Total
    ranking_venta = df_proveedores_grouped.copy()
    ranking_venta['ranking'] = ranking_venta['venta_total'].rank(
        ascending=False, method='min'
    ).astype(int)
    ranking_venta = ranking_venta.sort_values('venta_total', ascending=False)
    ranking_venta = ranking_venta[[
        'ranking', 'proveedor', 'venta_total', 'utilidad_total', 'cantidad_total',
        'rentabilidad_pct', 'articulos_unicos'
    ]]
    
    # Ranking 3: Por Utilidad Total
    ranking_utilidad = df_proveedores_grouped.copy()
    ranking_utilidad['ranking'] = ranking_utilidad['utilidad_total'].rank(
        ascending=False, method='min'
    ).astype(int)
    ranking_utilidad = ranking_utilidad.sort_values('utilidad_total', ascending=False)
    ranking_utilidad = ranking_utilidad[[
        'ranking', 'proveedor', 'utilidad_total', 'venta_total', 'cantidad_total',
        'rentabilidad_pct', 'articulos_unicos'
    ]]
    
    t1 = time.time()
    print_time("Rankings creados", t1-t0)
    
    # ========================================================================
    # 7. DETALLE POR ARTÍCULO
    # ========================================================================
    print("\n  Creando detalle por artículo...")
    t0 = time.time()
    
    # Columnas a incluir en el detalle
    columnas_detalle = ['proveedor', 'idarticulo']
    
    # Agregar columnas opcionales si existen
    if 'descripcion' in df_merged.columns:
        columnas_detalle.append('descripcion')
    if 'familia' in df_merged.columns:
        columnas_detalle.append('familia')
    if 'subfamilia' in df_merged.columns:
        columnas_detalle.append('subfamilia')
    
    # Agrupar por proveedor y artículo
    df_detalle = df_merged.groupby(columnas_detalle).agg({
        'cantidad_vendida': 'sum',
        'venta_total': 'sum',
        'costo_total': 'sum',
        'utilidad_total': 'sum'
    }).reset_index()
    
    # Renombrar la columna para mantener consistencia
    df_detalle.rename(columns={'cantidad_vendida': 'cantidad_total'}, inplace=True)
    
    # Calcular rentabilidad por artículo
    df_detalle['rentabilidad_pct'] = (
        df_detalle['utilidad_total'] / df_detalle['costo_total']
    ).round(4)
    
    # Detalle 1: Ordenado por Cantidad
    # Agregar orden del ranking de cantidad
    detalle_cantidad = df_detalle.copy()
    detalle_cantidad = detalle_cantidad.merge(
        ranking_cantidad[['proveedor', 'ranking']], 
        on='proveedor', 
        how='left'
    )
    detalle_cantidad = detalle_cantidad.sort_values(
        ['ranking', 'cantidad_total'], 
        ascending=[True, False]
    )
    detalle_cantidad = detalle_cantidad.drop(columns=['ranking'])
    
    # Detalle 2: Ordenado por Venta
    # Agregar orden del ranking de venta
    detalle_venta = df_detalle.copy()
    detalle_venta = detalle_venta.merge(
        ranking_venta[['proveedor', 'ranking']], 
        on='proveedor', 
        how='left'
    )
    detalle_venta = detalle_venta.sort_values(
        ['ranking', 'venta_total'], 
        ascending=[True, False]
    )
    detalle_venta = detalle_venta.drop(columns=['ranking'])
    
    # Detalle 3: Ordenado por Utilidad
    # Agregar orden del ranking de utilidad
    detalle_utilidad = df_detalle.copy()
    detalle_utilidad = detalle_utilidad.merge(
        ranking_utilidad[['proveedor', 'ranking']], 
        on='proveedor', 
        how='left'
    )
    detalle_utilidad = detalle_utilidad.sort_values(
        ['ranking', 'utilidad_total'], 
        ascending=[True, False]
    )
    detalle_utilidad = detalle_utilidad.drop(columns=['ranking'])
    
    t1 = time.time()
    print_time("Detalle por artículo creado", t1-t0)
    print_info("Registros en detalle", f"{len(df_detalle):,}")
    
    # ========================================================================
    # 8. LIMPIAR CARACTERES ILEGALES
    # ========================================================================
    print_subheader("🧹 LIMPIEZA DE DATOS")
    
    print("  Limpiando caracteres ilegales para Excel...")
    t0 = time.time()
    
    # Limpiar todos los DataFrames antes de exportar
    ranking_cantidad_clean = limpiar_dataframe_para_excel(ranking_cantidad)
    ranking_venta_clean = limpiar_dataframe_para_excel(ranking_venta)
    ranking_utilidad_clean = limpiar_dataframe_para_excel(ranking_utilidad)
    detalle_cantidad_clean = limpiar_dataframe_para_excel(detalle_cantidad)
    detalle_venta_clean = limpiar_dataframe_para_excel(detalle_venta)
    detalle_utilidad_clean = limpiar_dataframe_para_excel(detalle_utilidad)
    
    t1 = time.time()
    print_time("Limpieza completada", t1-t0)
    
    # ========================================================================
    # 9. GENERAR ARCHIVO EXCEL EN MEMORIA
    # ========================================================================
    print_subheader("💾 GENERACIÓN DE EXCEL")
    
    print("  Generando archivo Excel en memoria...")
    t0 = time.time()
    
    # Crear BytesIO para archivo en memoria
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Rankings de proveedores
        ranking_cantidad_clean.to_excel(writer, sheet_name='Ranking_Cantidad', index=False)
        ranking_venta_clean.to_excel(writer, sheet_name='Ranking_Venta', index=False)
        ranking_utilidad_clean.to_excel(writer, sheet_name='Ranking_Utilidad', index=False)
        
        # Detalles por artículo
        detalle_cantidad_clean.to_excel(writer, sheet_name='Detalle_por_Cantidad', index=False)
        detalle_venta_clean.to_excel(writer, sheet_name='Detalle_por_Venta', index=False)
        detalle_utilidad_clean.to_excel(writer, sheet_name='Detalle_por_Utilidad', index=False)
    
    t1 = time.time()
    print_time("Archivo Excel generado", t1-t0)
    
    # ========================================================================
    # 10. APLICAR FORMATO A TODAS LAS HOJAS
    # ========================================================================
    print_subheader("✨ APLICANDO FORMATO")
    
    # Reabrir el archivo desde BytesIO para aplicar formato
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        formatear_hoja(ws, sheet_name)
    
    # Guardar cambios en BytesIO
    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    
    # ========================================================================
    # 11. RESUMEN FINAL
    # ========================================================================
    print_subheader("📈 RESUMEN DE RESULTADOS")
    
    print(f"\n  {'Ranking por Cantidad':.<60}")
    print_info("Top 1 Proveedor", ranking_cantidad.iloc[0]['proveedor'].strip())
    print_info("Cantidad Total", f"{ranking_cantidad.iloc[0]['cantidad_total']:,.0f}")
    print_info("Rentabilidad", f"{ranking_cantidad.iloc[0]['rentabilidad_pct']:.2%}")
    
    print(f"\n  {'Ranking por Venta':.<60}")
    print_info("Top 1 Proveedor", ranking_venta.iloc[0]['proveedor'].strip())
    print_info("Venta Total", f"${ranking_venta.iloc[0]['venta_total']:,.2f}")
    print_info("Rentabilidad", f"{ranking_venta.iloc[0]['rentabilidad_pct']:.2%}")
    
    print(f"\n  {'Ranking por Utilidad':.<60}")
    print_info("Top 1 Proveedor", ranking_utilidad.iloc[0]['proveedor'].strip())
    print_info("Utilidad Total", f"${ranking_utilidad.iloc[0]['utilidad_total']:,.2f}")
    print_info("Rentabilidad", f"{ranking_utilidad.iloc[0]['rentabilidad_pct']:.2%}")
    
    # ========================================================================
    # TIEMPO TOTAL
    # ========================================================================
    tiempo_total_fin = time.time()
    tiempo_total = tiempo_total_fin - tiempo_total_inicio
    
    print_header("✅ PROCESO COMPLETADO EXITOSAMENTE")
    print_time("Tiempo total de ejecución", tiempo_total)
    print(f"\n  {'='*80}\n")
    
    return output_final