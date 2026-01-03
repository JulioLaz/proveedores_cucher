"""
═══════════════════════════════════════════════════════════════════════════════
    TAB 5: PREDICCIÓN Y PRESUPUESTO DE REABASTECIMIENTO
    
    Sistema de predicción basado en bloques de 7 días
    Adaptado para trabajar con BigQuery y Streamlit
    
    Autor: Julio Lazzaroni
    Fecha: Diciembre 2025
    
    VERSIÓN 2.0:
    - Calcula SIEMPRE 1, 2, 3, 4 semanas
    - Sin selector de semanas
    - Gráficas separadas por región (Chaco y Corrientes)
═══════════════════════════════════════════════════════════════════════════════
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from google.cloud import bigquery
import plotly.graph_objects as go

from utils.config import ID_LIST_SALTA, SALTA_REFRESCOS_ID, NOMBRES_UNIFICADOS

warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE CARGA DE DATOS DESDE BIGQUERY
# ═══════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=3600, show_spinner="Cargando datos de presupuesto...")
def cargar_datos_presupuesto_bq(credentials_path, project_id, id_list):
    """
    Carga datos del presupuesto desde BigQuery filtrados por IDs
    """
    try:
        client = bigquery.Client.from_service_account_json(credentials_path)
        
        id_str = ','.join(map(str, id_list))
        
        query = f"""
        SELECT *
        FROM `{project_id}.presupuesto.result_final_alert_all`
        WHERE idarticuloalfa IN ({id_str})
        """
        
        df = client.query(query).to_dataframe()
        
        st.success(f"✓ Presupuesto cargado: {len(df):,} registros")
        return df
    
    except Exception as e:
        st.error(f"Error al cargar presupuesto: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=3600, show_spinner="Cargando tickets históricos...")
def cargar_datos_tickets_bq(credentials_path, project_id, bigquery_table, id_list, fecha_desde):
    """
    Carga datos de tickets desde BigQuery filtrados por IDs y fecha
    """
    try:
        client = bigquery.Client.from_service_account_json(credentials_path)
        
        id_str = ','.join(map(str, id_list))
        
        query = f"""
        SELECT 
            fecha_comprobante,
            idartalfa,
            sucursal,
            cantidad_total as cantidad
        FROM `{project_id}.{bigquery_table}`
        WHERE idartalfa IN ({id_str})
        AND DATE(fecha_comprobante) >= '{fecha_desde}'
        ORDER BY fecha_comprobante
        """
        
        df = client.query(query).to_dataframe()
        
        if df.empty:
            st.warning("No se encontraron tickets para los artículos seleccionados")
            return pd.DataFrame()
        
        # Procesar fechas
        df['fecha_comprobante'] = pd.to_datetime(df['fecha_comprobante'])
        df['fecha'] = df['fecha_comprobante'].dt.date
        df['idartalfa'] = df['idartalfa'].astype(str)
        
        st.success(f"✓ Tickets cargados: {len(df):,} registros")
        st.info(f"Rango: {df['fecha'].min()} → {df['fecha'].max()}")
        
        return df
    
    except Exception as e:
        st.error(f"Error al cargar tickets: {e}")
        return pd.DataFrame()


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE TRANSFORMACIÓN: BLOQUES DE 7 DÍAS
# ═══════════════════════════════════════════════════════════════════════════════

def detectar_ultima_fecha_datos(df_tickets):
    """Detecta la última fecha con datos disponibles"""
    ultima_fecha = df_tickets['fecha'].max()
    return ultima_fecha


def clasificar_sucursal(sucursal):
    """
    Clasifica sucursal en Chaco, Corrientes o Formosa
    Formosa se mantiene separada para poder filtrarla opcionalmente
    """
    sucursal_str = str(sucursal).strip().lower()
    
    # CHACO: Hiper + almacenes (express, tirol, central)
    if sucursal_str in ['hiper', 'express', 'tirol', 'central', '6', '7', '8']:
        return 'chaco'
    
    # FORMOSA: Separada para filtrado condicional
    elif sucursal_str in ['formosa']:
        return 'formosa'
    
    # CORRIENTES
    elif sucursal_str in ['corrientes', '1', '3', '4', '5']:
        return 'corrientes'
    
    else:
        return 'otra'


def crear_bloques_7_dias(df_tickets, fecha_referencia):
    """Crea bloques de 7 días HACIA ATRÁS desde la fecha de referencia"""
    df = df_tickets.copy()
    
    if isinstance(fecha_referencia, pd.Timestamp):
        fecha_referencia = fecha_referencia.date()
    
    df['dias_desde_ref'] = df['fecha'].apply(lambda x: (fecha_referencia - x).days)
    df['bloque_7dias'] = (df['dias_desde_ref'] // 7) + 1
    df = df[df['bloque_7dias'] > 0].copy()
    df['region'] = df['sucursal'].apply(clasificar_sucursal)
    
    return df


def calcular_ventas_por_bloques(df_con_bloques, cantidad_bloques_historico=12):
    """Calcula las ventas totales por bloques de 7 días"""
    bloques_validos = sorted(df_con_bloques['bloque_7dias'].unique())[:cantidad_bloques_historico]
    df_reciente = df_con_bloques[df_con_bloques['bloque_7dias'].isin(bloques_validos)].copy()
    
    # Ventas totales
    ventas_totales = df_reciente.groupby(
        ['idartalfa', 'bloque_7dias']
    )['cantidad'].sum().reset_index()
    ventas_totales.rename(columns={'cantidad': 'cantidad_bloque'}, inplace=True)
    
    # Ventas por región
    ventas_por_region = df_reciente.groupby(
        ['idartalfa', 'bloque_7dias', 'region']
    )['cantidad'].sum().reset_index()
    ventas_por_region.rename(columns={'cantidad': 'cantidad_bloque_region'}, inplace=True)
    
    return ventas_totales, ventas_por_region


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE PREDICCIÓN
# ═══════════════════════════════════════════════════════════════════════════════

def calcular_demanda_potencial_bloques(ventas_totales, n_semanas=1):
    """Calcula la demanda potencial para N bloques de 7 días"""
    resultados = []
    
    for id_art, grupo in ventas_totales.groupby('idartalfa'):
        grupo_sorted = grupo.sort_values('bloque_7dias')
        
        if len(grupo_sorted) < 2:
            demanda_base = grupo_sorted['cantidad_bloque'].iloc[0] if len(grupo_sorted) > 0 else 0
        else:
            ultimos_bloques = grupo_sorted.head(min(4, len(grupo_sorted)))
            demanda_base = ultimos_bloques['cantidad_bloque'].mean()
            
            if len(grupo_sorted) >= 3:
                ultimos_3 = grupo_sorted.head(3)['cantidad_bloque'].values
                promedio_historico = grupo_sorted.tail(-3)['cantidad_bloque'].mean() if len(grupo_sorted) > 3 else demanda_base
                
                promedio_reciente = ultimos_3.mean()
                if promedio_historico > 0:
                    cambio_pct = (promedio_reciente - promedio_historico) / promedio_historico
                    
                    if cambio_pct > 0.15:
                        demanda_base *= 1.10
                    elif cambio_pct < -0.15:
                        demanda_base *= 0.95
        
        demanda_n_semanas = demanda_base * n_semanas
        
        resultados.append({
            'idartalfa': id_art,
            f'demanda_potencial_{n_semanas}_bloque{"s" if n_semanas > 1 else ""}': demanda_n_semanas
        })
    
    return pd.DataFrame(resultados)


def calcular_demanda_por_region(ventas_por_region, demanda_total_df, n_semanas=1):
    """Calcula demanda potencial por región"""
    if n_semanas == 1:
        col_demanda = 'demanda_potencial_proxima_semana'
    else:
        col_demanda = f'demanda_potencial_proximas_{n_semanas}_semanas'
    
    demanda_dict = demanda_total_df.set_index('idartalfa')[col_demanda].to_dict()
    
    resultados = []
    
    for id_art, grupo in ventas_por_region.groupby('idartalfa'):
        ventas_region = {}
        
        for region in ['chaco', 'corrientes']:
            grupo_region = grupo[grupo['region'] == region]
            
            if len(grupo_region) == 0:
                ventas_region[region] = 0
            else:
                grupo_sorted = grupo_region.sort_values('bloque_7dias')
                ultimos_bloques = grupo_sorted.head(min(4, len(grupo_sorted)))
                ventas_region[region] = ultimos_bloques['cantidad_bloque_region'].mean()
        
        total_historico = ventas_region['chaco'] + ventas_region['corrientes']
        
        if total_historico > 0:
            cor_perc_decimal = ventas_region['corrientes'] / total_historico
            chaco_perc_decimal = ventas_region['chaco'] / total_historico
        else:
            cor_perc_decimal = 0.5
            chaco_perc_decimal = 0.5
        
        demanda_total = demanda_dict.get(id_art, 0)
        
        resultados.append({
            'idartalfa': id_art,
            'chaco_cantidad': demanda_total * chaco_perc_decimal,
            'corr_cantidad': demanda_total * cor_perc_decimal,
            'cor_%': cor_perc_decimal * 100,
            'chaco_perc': chaco_perc_decimal
        })
    
    return pd.DataFrame(resultados)


def generar_columnas_demanda(ventas_totales, ventas_por_region):
    """Genera SIEMPRE las columnas de demanda para 1, 2, 3, 4 semanas"""
    
    # SIEMPRE calcular 1 semana
    df_1_semana = calcular_demanda_potencial_bloques(ventas_totales, n_semanas=1)
    df_1_semana.rename(columns={'demanda_potencial_1_bloque': 'demanda_potencial_proxima_semana'}, inplace=True)
    
    # SIEMPRE calcular 2 semanas
    df_2_semanas = calcular_demanda_potencial_bloques(ventas_totales, n_semanas=2)
    df_2_semanas = df_2_semanas[['idartalfa', 'demanda_potencial_2_bloques']]
    df_2_semanas.rename(columns={'demanda_potencial_2_bloques': 'demanda_potencial_proximas_2_semanas'}, inplace=True)
    
    # SIEMPRE calcular 3 semanas
    df_3_semanas = calcular_demanda_potencial_bloques(ventas_totales, n_semanas=3)
    df_3_semanas = df_3_semanas[['idartalfa', 'demanda_potencial_3_bloques']]
    df_3_semanas.rename(columns={'demanda_potencial_3_bloques': 'demanda_potencial_proximas_3_semanas'}, inplace=True)
    
    # SIEMPRE calcular 4 semanas
    df_4_semanas = calcular_demanda_potencial_bloques(ventas_totales, n_semanas=4)
    df_4_semanas = df_4_semanas[['idartalfa', 'demanda_potencial_4_bloques']]
    df_4_semanas.rename(columns={'demanda_potencial_4_bloques': 'demanda_potencial_proximas_4_semanas'}, inplace=True)
    
    # SIEMPRE calcular 30 días
    df_30_dias = calcular_demanda_potencial_bloques(ventas_totales, n_semanas=4.3)
    df_30_dias = df_30_dias[['idartalfa', 'demanda_potencial_4.3_bloques']]
    df_30_dias.rename(columns={'demanda_potencial_4.3_bloques': 'demanda_potencial_30_dias'}, inplace=True)
    
    # Merge todos
    df_resultado = df_1_semana.merge(df_2_semanas, on='idartalfa', how='left')
    df_resultado = df_resultado.merge(df_3_semanas, on='idartalfa', how='left')
    df_resultado = df_resultado.merge(df_4_semanas, on='idartalfa', how='left')
    df_resultado = df_resultado.merge(df_30_dias, on='idartalfa', how='left')
    
    # Calcular distribución por región (usar 4 semanas como referencia)
    df_distribucion = calcular_demanda_por_region(ventas_por_region, df_resultado, n_semanas=4)
    df_resultado = df_resultado.merge(df_distribucion, on='idartalfa', how='left')
    
    return df_resultado


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE CÁLCULO DE PRESUPUESTO
# ═══════════════════════════════════════════════════════════════════════════════

def agregar_info_productos(df_demanda, df_presupuesto):
    """Agrega información de productos desde presupuesto"""
    columnas_necesarias = [
        'idarticuloalfa', 'idarticulo', 'idproveedor', 'proveedor', 'familia', 'subfamilia',
        'descripcion', 'costo_unit', 'uxb',
        'stk_corrientes', 'stk_express', 'stk_hiper', 'stk_TIROL', 'stk_central'
    ]
    
    columnas_existentes = [col for col in columnas_necesarias if col in df_presupuesto.columns]
    
    df_info = df_presupuesto[columnas_existentes].copy()
    df_info = df_info.drop_duplicates(subset=['idarticuloalfa'], keep='first')
    df_info.rename(columns={'costo_unit': 'costo_unitario'}, inplace=True)
    
    for col in ['idarticulo', 'idproveedor', 'proveedor', 'familia', 'subfamilia']:
        if col not in df_info.columns:
            df_info[col] = ''
    
    for col in ['stk_corrientes', 'stk_express', 'stk_hiper', 'stk_TIROL', 'stk_central']:
        if col not in df_info.columns:
            df_info[col] = 0
    
    if 'uxb' not in df_info.columns:
        df_info['uxb'] = 1
    
    df_info['STK_CHACO'] = df_info['stk_express'] + df_info['stk_hiper'] + df_info['stk_TIROL'] + df_info['stk_central']
    df_info['STK_TOTAL'] = df_info['stk_corrientes'] + df_info['stk_express'] + df_info['stk_hiper'] + df_info['stk_TIROL'] + df_info['stk_central']
    
    df_demanda['idartalfa'] = df_demanda['idartalfa'].astype(int)
    df_info['idarticuloalfa'] = df_info['idarticuloalfa'].astype(int)
    
    df_final = df_demanda.merge(df_info, left_on='idartalfa', right_on='idarticuloalfa', how='left')
    df_final.drop('idarticuloalfa', axis=1, inplace=True)
    
    columnas_stock = ['stk_corrientes', 'stk_express', 'stk_hiper', 'stk_TIROL', 
                      'stk_central', 'STK_CHACO', 'STK_TOTAL']
    for col in columnas_stock:
        if col in df_final.columns:
            df_final[col] = df_final[col].fillna(0)
    
    return df_final


def generar_columnas_ventas_bloques(df_con_bloques, df_final, cantidad_bloques=4):
    """Genera columnas de ventas por bloques"""
    if df_con_bloques.empty:
        return df_final, []
    
    df_con_bloques['idartalfa'] = df_con_bloques['idartalfa'].astype(int)
    df_final['idartalfa'] = df_final['idartalfa'].astype(int)
    
    bloques_unicos = sorted(df_con_bloques['bloque_7dias'].unique())[:cantidad_bloques]
    bloques_ordenados = list(reversed(bloques_unicos))
    
    nombres_bloques = []
    for bloque in bloques_ordenados:
        df_bloque = df_con_bloques[df_con_bloques['bloque_7dias'] == bloque]
        fecha_min = df_bloque['fecha'].min()
        fecha_max = df_bloque['fecha'].max()
        nombre = f"sem_{fecha_min.strftime('%d%b')}_{fecha_max.strftime('%d%b')}".lower()
        nombres_bloques.append(nombre)
    
    for nombre in nombres_bloques:
        df_final[nombre] = 0
        df_final[f"{nombre}_chaco"] = 0
        df_final[f"{nombre}_corr"] = 0
    
    for idx, bloque in enumerate(bloques_ordenados):
        nombre = nombres_bloques[idx]
        df_bloque = df_con_bloques[df_con_bloques['bloque_7dias'] == bloque].copy()
        
        # Totales
        ventas_total = df_bloque.groupby('idartalfa')['cantidad'].sum().to_dict()
        for idartalfa, cantidad in ventas_total.items():
            df_final.loc[df_final['idartalfa'] == idartalfa, nombre] = int(cantidad)
        
        # Chaco
        df_chaco = df_bloque[df_bloque['region'] == 'chaco']
        ventas_chaco = df_chaco.groupby('idartalfa')['cantidad'].sum().to_dict()
        for idartalfa, cantidad in ventas_chaco.items():
            df_final.loc[df_final['idartalfa'] == idartalfa, f"{nombre}_chaco"] = int(cantidad)
        
        # Corrientes
        df_corr = df_bloque[df_bloque['region'] == 'corrientes']
        ventas_corr = df_corr.groupby('idartalfa')['cantidad'].sum().to_dict()
        for idartalfa, cantidad in ventas_corr.items():
            df_final.loc[df_final['idartalfa'] == idartalfa, f"{nombre}_corr"] = int(cantidad)
    
    # ═══════════════════════════════════════════════════════════════════════
    # GENERAR COLUMNA TOTAL (suma de todas las semanas)
    # ═══════════════════════════════════════════════════════════════════════
    if len(nombres_bloques) >= 4:
        fecha_inicio = nombres_bloques[0].split('_')[1]
        fecha_fin = nombres_bloques[-1].split('_')[-1]
        nombre_total = f"TOTAL_{fecha_inicio}_{fecha_fin}"
        
        df_final[nombre_total] = 0
        for nombre_bloque in nombres_bloques:
            if nombre_bloque in df_final.columns:
                df_final[nombre_total] += df_final[nombre_bloque]
        
        print(f"✓ Columna {nombre_total} creada: suma de {len(nombres_bloques)} semanas")
    
    return df_final, nombres_bloques


def calcular_presupuesto(df_final):
    """Calcula presupuesto para 1, 2, 3 y 4 semanas"""
    
    # ═══ 1 SEMANA ═══
    df_final['corr_abastecer_1sem'] = (df_final['demanda_potencial_proxima_semana'] * df_final['cor_%'] / 100) - df_final['stk_corrientes']
    df_final['corr_abastecer_1sem'] = df_final['corr_abastecer_1sem'].clip(lower=0)
    
    df_final['chaco_abastecer_1sem'] = (df_final['demanda_potencial_proxima_semana'] * df_final['chaco_perc']) - df_final['STK_CHACO']
    df_final['chaco_abastecer_1sem'] = df_final['chaco_abastecer_1sem'].clip(lower=0)
    
    df_final['presupuesto_corrientes_1sem'] = df_final['corr_abastecer_1sem'] * df_final['costo_unitario']
    df_final['presupuesto_chaco_1sem'] = df_final['chaco_abastecer_1sem'] * df_final['costo_unitario']
    df_final['presupuesto_total_1sem'] = df_final['presupuesto_corrientes_1sem'] + df_final['presupuesto_chaco_1sem']
    
    # ═══ 2 SEMANAS ═══
    df_final['corr_abastecer_2sem'] = (df_final['demanda_potencial_proximas_2_semanas'] * df_final['cor_%'] / 100) - df_final['stk_corrientes']
    df_final['corr_abastecer_2sem'] = df_final['corr_abastecer_2sem'].clip(lower=0)
    
    df_final['chaco_abastecer_2sem'] = (df_final['demanda_potencial_proximas_2_semanas'] * df_final['chaco_perc']) - df_final['STK_CHACO']
    df_final['chaco_abastecer_2sem'] = df_final['chaco_abastecer_2sem'].clip(lower=0)
    
    df_final['presupuesto_corrientes_2sem'] = df_final['corr_abastecer_2sem'] * df_final['costo_unitario']
    df_final['presupuesto_chaco_2sem'] = df_final['chaco_abastecer_2sem'] * df_final['costo_unitario']
    df_final['presupuesto_total_2sem'] = df_final['presupuesto_corrientes_2sem'] + df_final['presupuesto_chaco_2sem']
    
    # ═══ 3 SEMANAS ═══
    df_final['corr_abastecer_3sem'] = (df_final['demanda_potencial_proximas_3_semanas'] * df_final['cor_%'] / 100) - df_final['stk_corrientes']
    df_final['corr_abastecer_3sem'] = df_final['corr_abastecer_3sem'].clip(lower=0)
    
    df_final['chaco_abastecer_3sem'] = (df_final['demanda_potencial_proximas_3_semanas'] * df_final['chaco_perc']) - df_final['STK_CHACO']
    df_final['chaco_abastecer_3sem'] = df_final['chaco_abastecer_3sem'].clip(lower=0)
    
    df_final['presupuesto_corrientes_3sem'] = df_final['corr_abastecer_3sem'] * df_final['costo_unitario']
    df_final['presupuesto_chaco_3sem'] = df_final['chaco_abastecer_3sem'] * df_final['costo_unitario']
    df_final['presupuesto_total_3sem'] = df_final['presupuesto_corrientes_3sem'] + df_final['presupuesto_chaco_3sem']
    
    # ═══ 4 SEMANAS ═══
    df_final['corr_abastecer_4sem'] = (df_final['demanda_potencial_proximas_4_semanas'] * df_final['cor_%'] / 100) - df_final['stk_corrientes']
    df_final['corr_abastecer_4sem'] = df_final['corr_abastecer_4sem'].clip(lower=0)
    
    df_final['chaco_abastecer_4sem'] = (df_final['demanda_potencial_proximas_4_semanas'] * df_final['chaco_perc']) - df_final['STK_CHACO']
    df_final['chaco_abastecer_4sem'] = df_final['chaco_abastecer_4sem'].clip(lower=0)
    
    df_final['presupuesto_corrientes_4sem'] = df_final['corr_abastecer_4sem'] * df_final['costo_unitario']
    df_final['presupuesto_chaco_4sem'] = df_final['chaco_abastecer_4sem'] * df_final['costo_unitario']
    df_final['presupuesto_total_4sem'] = df_final['presupuesto_corrientes_4sem'] + df_final['presupuesto_chaco_4sem']
    
    return df_final


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE EXPORTACIÓN A EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def formatear_hoja_excel(ws, destacar_columna=None):
    """Aplica formato profesional a una hoja Excel"""
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Calibri', size=10)
    data_font_bold = Font(name='Calibri', size=10, bold=True)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    highlight_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    
    ws.row_dimensions[1].height = 25
    
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        es_par = row_idx % 2 == 0
        
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value
            
            cell.fill = light_fill if es_par else white_fill
            cell.font = data_font
            cell.border = thin_border
            
            if destacar_columna and col_name == destacar_columna:
                cell.fill = highlight_fill
            
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if 'presupuesto' in str(col_name).lower():
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = data_alignment
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = max(max_length + 5, 12)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = 'A2'


def crear_libro_excel_formateado(df_final, nombres_bloques, proveedor_nombre):
    """Crea el libro Excel con formato profesional y retorna bytes"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # HOJA GENERAL
    ws_general = wb.create_sheet("GENERAL")
    
    columnas_mostrar = [
        'idartalfa', 'idarticulo', 'idproveedor', 'proveedor', 'familia', 'subfamilia',
        'descripcion', 'uxb', 'costo_unitario',
        'stk_corrientes', 'stk_express', 'stk_hiper', 'stk_central', 'stk_TIROL',
        'STK_CHACO', 'STK_TOTAL'
    ]
    
    for nombre_bloque in nombres_bloques:
        columnas_mostrar.append(nombre_bloque)
    
    for nombre_bloque in nombres_bloques:
        columnas_mostrar.append(f"{nombre_bloque}_chaco")
    for nombre_bloque in nombres_bloques:
        columnas_mostrar.append(f"{nombre_bloque}_corr")
    
    columnas_mostrar.extend([
        'demanda_potencial_proxima_semana',
        'demanda_potencial_proximas_2_semanas',
        'demanda_potencial_proximas_3_semanas',
        'demanda_potencial_proximas_4_semanas',
        'demanda_potencial_30_dias',
        'chaco_cantidad', 'corr_cantidad',
        'corr_abastecer_1sem', 'chaco_abastecer_1sem',
        'corr_abastecer_2sem', 'chaco_abastecer_2sem',
        'corr_abastecer_3sem', 'chaco_abastecer_3sem',
        'corr_abastecer_4sem', 'chaco_abastecer_4sem',
        'presupuesto_chaco_1sem', 'presupuesto_corrientes_1sem', 'presupuesto_total_1sem',
        'presupuesto_chaco_2sem', 'presupuesto_corrientes_2sem', 'presupuesto_total_2sem',
        'presupuesto_chaco_3sem', 'presupuesto_corrientes_3sem', 'presupuesto_total_3sem',
        'presupuesto_chaco_4sem', 'presupuesto_corrientes_4sem', 'presupuesto_total_4sem'
    ])
    
    columnas_existentes = [col for col in columnas_mostrar if col in df_final.columns]
    df_export = df_final[columnas_existentes].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
        ws_general.append(row)
    
    formatear_hoja_excel(ws_general)
    
    # HOJA CHACO
    ws_chaco = wb.create_sheet("CHACO")
    df_chaco = df_final[(df_final['chaco_abastecer_1sem'] > 0) | 
                        (df_final['chaco_abastecer_2sem'] > 0) |
                        (df_final['chaco_abastecer_3sem'] > 0) |
                        (df_final['chaco_abastecer_4sem'] > 0)].copy()
    
    if len(nombres_bloques) >= 1:
        ultima_semana_chaco = f"{nombres_bloques[-1]}_chaco"
        df_chaco['ultimos_7_dias'] = df_chaco[ultima_semana_chaco] if ultima_semana_chaco in df_chaco.columns else 0
    
    if len(nombres_bloques) >= 2:
        ultima_sem_chaco = f"{nombres_bloques[-1]}_chaco"
        anteultima_sem_chaco = f"{nombres_bloques[-2]}_chaco"
        df_chaco['ultimos_14_dias'] = (df_chaco[ultima_sem_chaco] if ultima_sem_chaco in df_chaco.columns else 0) + \
                                       (df_chaco[anteultima_sem_chaco] if anteultima_sem_chaco in df_chaco.columns else 0)
    
    columnas_chaco = ['idartalfa', 'descripcion', 'uxb', 'ultimos_7_dias', 'ultimos_14_dias', 
                      'chaco_abastecer_1sem', 'chaco_abastecer_2sem', 'chaco_abastecer_3sem', 'chaco_abastecer_4sem',
                      'costo_unitario', 
                      'presupuesto_chaco_1sem', 'presupuesto_chaco_2sem', 
                      'presupuesto_chaco_3sem', 'presupuesto_chaco_4sem']
    columnas_chaco_existentes = [col for col in columnas_chaco if col in df_chaco.columns]
    df_chaco_export = df_chaco[columnas_chaco_existentes].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(df_chaco_export, index=False, header=True), 1):
        ws_chaco.append(row)
    
    formatear_hoja_excel(ws_chaco)
    
    # HOJA CORRIENTES
    ws_corr = wb.create_sheet("CORRIENTES")
    df_corr = df_final[(df_final['corr_abastecer_1sem'] > 0) | 
                       (df_final['corr_abastecer_2sem'] > 0) |
                       (df_final['corr_abastecer_3sem'] > 0) |
                       (df_final['corr_abastecer_4sem'] > 0)].copy()
    
    if len(nombres_bloques) >= 1:
        ultima_semana_corr = f"{nombres_bloques[-1]}_corr"
        df_corr['ultimos_7_dias'] = df_corr[ultima_semana_corr] if ultima_semana_corr in df_corr.columns else 0
    
    if len(nombres_bloques) >= 2:
        ultima_sem_corr = f"{nombres_bloques[-1]}_corr"
        anteultima_sem_corr = f"{nombres_bloques[-2]}_corr"
        df_corr['ultimos_14_dias'] = (df_corr[ultima_sem_corr] if ultima_sem_corr in df_corr.columns else 0) + \
                                      (df_corr[anteultima_sem_corr] if anteultima_sem_corr in df_corr.columns else 0)
    
    columnas_corr = ['idartalfa', 'descripcion', 'uxb', 'ultimos_7_dias', 'ultimos_14_dias',
                     'corr_abastecer_1sem', 'corr_abastecer_2sem', 'corr_abastecer_3sem', 'corr_abastecer_4sem',
                     'costo_unitario', 
                     'presupuesto_corrientes_1sem', 'presupuesto_corrientes_2sem',
                     'presupuesto_corrientes_3sem', 'presupuesto_corrientes_4sem']
    columnas_corr_existentes = [col for col in columnas_corr if col in df_corr.columns]
    df_corr_export = df_corr[columnas_corr_existentes].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(df_corr_export, index=False, header=True), 1):
        ws_corr.append(row)
    
    formatear_hoja_excel(ws_corr)
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL DE LA TAB
# ═══════════════════════════════════════════════════════════════════════════════

def render_tab_prediccion_presupuesto(df_proveedores, config):
    """
    Renderiza la tab de Predicción y Presupuesto
    
    Args:
        df_proveedores: DataFrame con información de proveedores
        config: Diccionario con configuración (credentials, project_id, etc.)
    """
    
    st.markdown("""
        <div style="text-align: center; padding: 1.5rem; border: 2px solid #4CAF50; 
        border-radius: 10px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
        margin-bottom: 2rem;">
            <h1 style="color: white; margin: 0; font-size: 2.5rem;">
                🎯 Predicción y Presupuesto de Reabastecimiento
            </h1>
            <p style="color: white; margin-top: 0.5rem; font-size: 1.2rem;">
                Sistema de bloques de 7 días con cálculo automático 1-4 semanas
            </p>
        </div>
    """, unsafe_allow_html=True)
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SELECTOR DE PROVEEDOR
    # ═══════════════════════════════════════════════════════════════════════════
    
    # Crear lista de proveedores disponibles
    proveedores_disponibles = df_proveedores[['idproveedor', 'proveedor']].drop_duplicates()
    
    # Agregar SALTA REFRESCOS si no está
    if SALTA_REFRESCOS_ID not in proveedores_disponibles['idproveedor'].values:
        salta_row = pd.DataFrame({
            'idproveedor': [SALTA_REFRESCOS_ID],
            'proveedor': [NOMBRES_UNIFICADOS[SALTA_REFRESCOS_ID]]
        })
        proveedores_disponibles = pd.concat([salta_row, proveedores_disponibles], ignore_index=True)
    
    proveedores_disponibles = proveedores_disponibles.sort_values('proveedor')
    
    # Encontrar índice de SALTA REFRESCOS
    lista_proveedores = proveedores_disponibles['proveedor'].tolist()
    index_salta = lista_proveedores.index(NOMBRES_UNIFICADOS[SALTA_REFRESCOS_ID]) if NOMBRES_UNIFICADOS[SALTA_REFRESCOS_ID] in lista_proveedores else 0
    
    # Selectbox con SALTA por defecto
    proveedor_seleccionado = st.selectbox(
        "🏢 Seleccionar Proveedor:",
        options=lista_proveedores,
        index=index_salta
    )
    
    id_proveedor_seleccionado = proveedores_disponibles[
        proveedores_disponibles['proveedor'] == proveedor_seleccionado
    ]['idproveedor'].iloc[0]
    
    # ═══════════════════════════════════════════════════════════════════════════
    # CHECKBOX PARA INCLUIR FORMOSA
    # ═══════════════════════════════════════════════════════════════════════════
    
    incluir_formosa = st.checkbox(
        "Incluir sucursal Formosa en análisis de ventas",
        value=False,
        help="Formosa representa <5% de ventas. Por defecto está excluida para simplificar análisis."
    )
    
    if incluir_formosa:
        st.info("✓ Formosa incluida: se agrupará con región CHACO")
    else:
        st.info("✓ Formosa excluida: solo se analizarán Hiper (Chaco) y Corrientes")
    
    # Determinar IDs de artículos según proveedor
    if id_proveedor_seleccionado == SALTA_REFRESCOS_ID:
        ids_articulos = ID_LIST_SALTA
        st.info(f"📊 Proveedor virtual: {len(ids_articulos)} artículos agrupados")
    else:
        ids_articulos = df_proveedores[
            df_proveedores['idproveedor'] == id_proveedor_seleccionado
        ]['idarticulo'].unique().tolist()
        
        if len(ids_articulos) == 0:
            st.warning("⚠️ No se encontraron artículos para este proveedor")
            return
        
        st.info(f"📦 {len(ids_articulos)} artículos encontrados")
    
    # ═══════════════════════════════════════════════════════════════════════════
    # BOTÓN DE GENERAR ANÁLISIS
    # ═══════════════════════════════════════════════════════════════════════════
    
    if st.button("🚀 Generar Análisis y Presupuesto", type="primary", width='stretch'):
        
        with st.spinner("Procesando datos..."):
            
            # Cargar datos de presupuesto
            df_presupuesto = cargar_datos_presupuesto_bq(
                config['credentials_path'],
                config['project_id'],
                ids_articulos
            )
            
            if df_presupuesto.empty:
                st.error("❌ No se encontraron datos de presupuesto")
                return
            
            # Cargar tickets (últimos 90 días)
            fecha_desde = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
            
            df_tickets = cargar_datos_tickets_bq(
                config['credentials_path'],
                config['project_id'],
                config['bigquery_table'],
                ids_articulos,
                fecha_desde
            )
            
            if df_tickets.empty:
                st.error("❌ No se encontraron tickets históricos")
                return
            
            # Detectar última fecha
            ultima_fecha = detectar_ultima_fecha_datos(df_tickets)
            st.success(f"✓ Última fecha de datos: {ultima_fecha.strftime('%d-%b-%Y')}")
            
            # Crear bloques
            df_con_bloques = crear_bloques_7_dias(df_tickets, ultima_fecha)
            
            # ═══════════════════════════════════════════════════════════════════
            # FILTRAR O INCLUIR FORMOSA SEGÚN CHECKBOX
            # ═══════════════════════════════════════════════════════════════════
            if incluir_formosa:
                # Agrupar Formosa con Chaco
                df_con_bloques.loc[df_con_bloques['region'] == 'formosa', 'region'] = 'chaco'
                st.info("✓ Formosa agrupada con CHACO para el análisis")
            else:
                # Excluir Formosa del análisis
                registros_formosa = len(df_con_bloques[df_con_bloques['region'] == 'formosa'])
                df_con_bloques = df_con_bloques[df_con_bloques['region'] != 'formosa'].copy()
                if registros_formosa > 0:
                    st.info(f"✓ Formosa excluida: {registros_formosa:,} registros filtrados")
            
            # Calcular ventas por bloques
            ventas_totales, ventas_por_region = calcular_ventas_por_bloques(df_con_bloques)
            
            # Generar columnas de demanda (SIEMPRE 1, 2, 3, 4)
            df_demanda = generar_columnas_demanda(ventas_totales, ventas_por_region)
            
            # Agregar info de productos
            df_final = agregar_info_productos(df_demanda, df_presupuesto)
            
            # Generar columnas de ventas por bloques
            df_final, nombres_bloques = generar_columnas_ventas_bloques(df_con_bloques, df_final)
            
            # Calcular presupuesto (SIEMPRE 1, 2, 3, 4)
            df_final = calcular_presupuesto(df_final)
            
            # ═══════════════════════════════════════════════════════════════════
            # RESUMEN DE PRESUPUESTO (4 COLUMNAS)
            # ═══════════════════════════════════════════════════════════════════
            
            st.markdown("---")
            st.subheader("📊 Resumen de Presupuesto")
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown("### 📅 1 Semana")
                total_1sem = df_final['presupuesto_total_1sem'].sum()
                chaco_1sem = df_final['presupuesto_chaco_1sem'].sum()
                corr_1sem = df_final['presupuesto_corrientes_1sem'].sum()
                unidades_1sem = df_final['corr_abastecer_1sem'].sum() + df_final['chaco_abastecer_1sem'].sum()
                
                st.metric("💰 Total", f"${total_1sem:,.2f}")
                st.metric("🟧 Chaco", f"${chaco_1sem:,.2f}")
                st.metric("🟦 Corrientes", f"${corr_1sem:,.2f}")
                st.metric("📦 Unidades", f"{unidades_1sem:,.0f}")
            
            with col2:
                st.markdown("### 📅 2 Semanas")
                total_2sem = df_final['presupuesto_total_2sem'].sum()
                chaco_2sem = df_final['presupuesto_chaco_2sem'].sum()
                corr_2sem = df_final['presupuesto_corrientes_2sem'].sum()
                unidades_2sem = df_final['corr_abastecer_2sem'].sum() + df_final['chaco_abastecer_2sem'].sum()
                
                st.metric("💰 Total", f"${total_2sem:,.2f}")
                st.metric("🟧 Chaco", f"${chaco_2sem:,.2f}")
                st.metric("🟦 Corrientes", f"${corr_2sem:,.2f}")
                st.metric("📦 Unidades", f"{unidades_2sem:,.0f}")
            
            with col3:
                st.markdown("### 📅 3 Semanas")
                total_3sem = df_final['presupuesto_total_3sem'].sum()
                chaco_3sem = df_final['presupuesto_chaco_3sem'].sum()
                corr_3sem = df_final['presupuesto_corrientes_3sem'].sum()
                unidades_3sem = df_final['corr_abastecer_3sem'].sum() + df_final['chaco_abastecer_3sem'].sum()
                
                st.metric("💰 Total", f"${total_3sem:,.2f}")
                st.metric("🟧 Chaco", f"${chaco_3sem:,.2f}")
                st.metric("🟦 Corrientes", f"${corr_3sem:,.2f}")
                st.metric("📦 Unidades", f"{unidades_3sem:,.0f}")
            
            with col4:
                st.markdown("### 📅 4 Semanas")
                total_4sem = df_final['presupuesto_total_4sem'].sum()
                chaco_4sem = df_final['presupuesto_chaco_4sem'].sum()
                corr_4sem = df_final['presupuesto_corrientes_4sem'].sum()
                unidades_4sem = df_final['corr_abastecer_4sem'].sum() + df_final['chaco_abastecer_4sem'].sum()
                
                st.metric("💰 Total", f"${total_4sem:,.2f}")
                st.metric("🟧 Chaco", f"${chaco_4sem:,.2f}")
                st.metric("🟦 Corrientes", f"${corr_4sem:,.2f}")
                st.metric("📦 Unidades", f"{unidades_4sem:,.0f}")
            
            # ═══════════════════════════════════════════════════════════════════
            # MOSTRAR TABLAS
            # ═══════════════════════════════════════════════════════════════════
            
            st.markdown("---")
            st.subheader("📋 Tablas de Datos Detallados")
            
            # TABLA GENERAL (ancho completo)
            with st.expander("📊 Ver Tabla GENERAL", expanded=False):
                st.dataframe(df_final, width='stretch', height=400)
            
            # TABLAS CHACO y CORRIENTES en paralelo
            col_t1, col_t2 = st.columns(2)
            
            with col_t1:
                with st.expander("🟧 Ver Tabla CHACO", expanded=False):
                    df_chaco_tabla = df_final[(df_final['chaco_abastecer_1sem'] > 0) | 
                                              (df_final['chaco_abastecer_2sem'] > 0) |
                                              (df_final['chaco_abastecer_3sem'] > 0) |
                                              (df_final['chaco_abastecer_4sem'] > 0)].copy()
                    st.dataframe(df_chaco_tabla, width='stretch', height=400)
            
            with col_t2:
                with st.expander("🟦 Ver Tabla CORRIENTES", expanded=False):
                    df_corr_tabla = df_final[(df_final['corr_abastecer_1sem'] > 0) | 
                                             (df_final['corr_abastecer_2sem'] > 0) |
                                             (df_final['corr_abastecer_3sem'] > 0) |
                                             (df_final['corr_abastecer_4sem'] > 0)].copy()
                    st.dataframe(df_corr_tabla, width='stretch', height=400)
            
            # ═══════════════════════════════════════════════════════════════════
            # FUNCIONES AUXILIARES PARA GRÁFICOS
            # ═══════════════════════════════════════════════════════════════════
            
            def calcular_dias_cobertura(row, columna_venta_periodo, dias_periodo=28):
                """Calcula días de cobertura basado en venta del período"""
                if columna_venta_periodo in row and row[columna_venta_periodo] > 0:
                    venta_diaria = row[columna_venta_periodo] / dias_periodo
                    return row['STK_TOTAL'] / venta_diaria
                return 0
            
            def get_color_cobertura(dias):
                """Asigna color según días de cobertura"""
                if dias < 7:
                    return '#e74c3c'
                elif dias < 14:
                    return '#e67e22'
                elif dias < 21:
                    return '#f39c12'
                else:
                    return '#27ae60'
            
            # ═══════════════════════════════════════════════════════════════════
            # GRÁFICOS 1: TOTAL PERÍODO + COBERTURA
            # ═══════════════════════════════════════════════════════════════════
            
            st.markdown("---")
            st.subheader("📊 Análisis de Ventas Período Completo")
            
            # Identificar columna TOTAL
            columna_total = None
            for col in df_final.columns:
                if col.startswith('TOTAL_'):
                    columna_total = col
                    break
            
            if columna_total:
                # Top 20 por cantidad total del período
                df_top_periodo = df_final.nlargest(20, columna_total).copy()
                df_top_periodo = df_top_periodo.iloc[::-1]
                
                df_top_periodo['articulo_label'] = df_top_periodo.apply(
                    lambda row: f"{row['idartalfa']} - {row['descripcion'][:30]}..." 
                    if len(str(row['descripcion'])) > 30 
                    else f"{row['idartalfa']} - {row['descripcion']}", 
                    axis=1
                )
                
                df_top_periodo['dias_cobertura'] = df_top_periodo.apply(
                    lambda row: calcular_dias_cobertura(row, columna_total, dias_periodo=28), axis=1
                )
                df_top_periodo['cobertura_visual'] = df_top_periodo['dias_cobertura'].apply(lambda x: min(x, 31))
                df_top_periodo['cobertura_texto'] = df_top_periodo['dias_cobertura'].apply(
                    lambda x: f"{x:.0f}d" if x <= 31 else f"31d+"
                )
                df_top_periodo['color_cobertura'] = df_top_periodo['dias_cobertura'].apply(get_color_cobertura)
                
                col_g1, col_g2 = st.columns(2)
                
                with col_g1:
                    st.markdown(f"##### 📦 Top 20 por Cantidad Vendida ({columna_total.replace('TOTAL_', '').replace('_', ' al ')})")
                    
                    hover_text_cantidad = []
                    for idx, row in df_top_periodo.iterrows():
                        texto = f"<b>{row['articulo_label']}</b><br>"
                        texto += f"Cantidad Total: {int(row[columna_total]):,}<br>"
                        texto += f"Stock Total: {int(row['STK_TOTAL']):,}<br>"
                        texto += f"Cobertura: {row['dias_cobertura']:.0f} días"
                        hover_text_cantidad.append(texto)
                    
                    fig_cantidad = go.Figure()
                    
                    fig_cantidad.add_trace(go.Bar(
                        y=df_top_periodo['articulo_label'],
                        x=df_top_periodo[columna_total],
                        orientation='h',
                        text=df_top_periodo[columna_total].apply(lambda x: f"{int(x):,}"),
                        textposition='outside',
                        cliponaxis=False,
                        marker=dict(color='#3498db', line=dict(width=0)),
                        hovertemplate='%{customdata}<extra></extra>',
                        customdata=hover_text_cantidad
                    ))
                    
                    max_cantidad = df_top_periodo[columna_total].max()
                    
                    fig_cantidad.update_layout(
                        height=max(400, 20 * 25),
                        margin=dict(t=20, b=25, l=10, r=100),
                        xaxis=dict(visible=False, range=[0, max_cantidad * 1.2]),
                        yaxis=dict(visible=True, tickfont=dict(size=10)),
                        showlegend=False,
                        plot_bgcolor='white',
                        paper_bgcolor='white'
                    )
                    
                    st.plotly_chart(fig_cantidad, width='stretch')
                
                with col_g2:
                    st.markdown("##### ⏱️ Días de Cobertura (Cap: 31 días)")
                    
                    hover_text_cobertura = []
                    for idx, row in df_top_periodo.iterrows():
                        texto = f"<b>{row['articulo_label']}</b><br>"
                        texto += f"Cobertura: {row['dias_cobertura']:.0f} días<br>"
                        texto += f"Stock Total: {int(row['STK_TOTAL']):,}<br>"
                        texto += f"Venta Período: {int(row[columna_total]):,}"
                        hover_text_cobertura.append(texto)
                    
                    fig_cobertura = go.Figure()
                    
                    fig_cobertura.add_trace(go.Bar(
                        y=df_top_periodo['articulo_label'],
                        x=df_top_periodo['cobertura_visual'],
                        orientation='h',
                        text=df_top_periodo['cobertura_texto'],
                        textposition='outside',
                        cliponaxis=False,
                        marker=dict(color=df_top_periodo['color_cobertura'], line=dict(width=0)),
                        hovertemplate='%{customdata}<extra></extra>',
                        customdata=hover_text_cobertura
                    ))
                    
                    lineas_dias = [7, 14, 21, 28]
                    colores_lineas = ['#e74c3c', '#e67e22', '#f39c12', '#3498db']
                    
                    for dia, color in zip(lineas_dias, colores_lineas):
                        fig_cobertura.add_vline(
                            x=dia, line_dash="dash", line_color=color, line_width=1.5, opacity=0.6,
                            annotation_text=f"{dia}d", annotation_position="top",
                            annotation_font_size=9, annotation_font_color=color
                        )
                    
                    fig_cobertura.update_layout(
                        height=max(400, 20 * 25),
                        margin=dict(t=20, b=5, l=30, r=20),
                        xaxis=dict(
                            visible=True, range=[0, 33], tickmode='array',
                            tickvals=[0, 7, 14, 21, 28, 31],
                            ticktext=['0', '7', '14', '21', '28', '31+'],
                            tickfont=dict(size=9)
                        ),
                        yaxis=dict(visible=True, tickfont=dict(size=10)),
                        showlegend=False,
                        plot_bgcolor='white',
                        paper_bgcolor='white'
                    )
                    
                    st.plotly_chart(fig_cobertura, width='stretch')
            
            # ═══════════════════════════════════════════════════════════════════
            # GRÁFICOS 2: PRESUPUESTO CHACO 1 SEMANA + COBERTURA
            # ═══════════════════════════════════════════════════════════════════
            
            st.markdown("---")
            st.subheader("🟧 Análisis de Presupuesto CHACO - 1 Semana")
            
            df_top_chaco = df_final.nlargest(20, 'presupuesto_chaco_1sem').copy()
            df_top_chaco = df_top_chaco.iloc[::-1]
            
            df_top_chaco['articulo_label'] = df_top_chaco.apply(
                lambda row: f"{row['idartalfa']} - {row['descripcion'][:30]}..." 
                if len(str(row['descripcion'])) > 30 
                else f"{row['idartalfa']} - {row['descripcion']}", 
                axis=1
            )
            
            df_top_chaco['dias_cobertura'] = df_top_chaco.apply(
                lambda row: (row['STK_CHACO'] / (row['chaco_cantidad'] / 7)) 
                if row['chaco_cantidad'] > 0 else 0,
                axis=1
            )
            df_top_chaco['cobertura_visual'] = df_top_chaco['dias_cobertura'].apply(lambda x: min(x, 31))
            df_top_chaco['cobertura_texto'] = df_top_chaco['dias_cobertura'].apply(
                lambda x: f"{x:.0f}d" if x <= 31 else f"31d+"
            )
            df_top_chaco['color_cobertura'] = df_top_chaco['dias_cobertura'].apply(get_color_cobertura)
            
            col_g3, col_g4 = st.columns(2)
            
            with col_g3:
                st.markdown("##### 💵 Top 20 por Presupuesto CHACO y unidades a reabastecer")
                
                # df_top_chaco['Presupuesto_M'] = df_top_chaco['presupuesto_chaco_1sem'] / 1_000_000
                # df_top_chaco['Texto'] = df_top_chaco['presupuesto_chaco_1sem'].apply(
                #     lambda x: f"${x:,.0f}".replace(",", ".")
                # )
                
                df_top_chaco['Presupuesto_M'] = df_top_chaco['presupuesto_chaco_1sem'] / 1_000_000

                # Construir texto con presupuesto y unidades a reabastecer
                df_top_chaco['Texto'] = df_top_chaco.apply(
                    lambda row: f"${row['presupuesto_chaco_1sem']:,.0f}".replace(",", ".") + 
                                f" | {int(row['chaco_abastecer_1sem']):,} u".replace(",", "."),
                    axis=1
                )


                hover_text_pres = []
                for idx, row in df_top_chaco.iterrows():
                    texto = f"<b>{row['articulo_label']}</b><br>"
                    texto += f"Presupuesto Chaco: ${row['presupuesto_chaco_1sem']:,.2f}<br>"
                    texto += f"A reabastecer: {int(row['chaco_abastecer_1sem']):,} uds<br>"
                    texto += f"Stock CHACO actual: {int(row['STK_CHACO']):,}"
                    hover_text_pres.append(texto)
                
                fig_presupuesto = go.Figure()
                
                fig_presupuesto.add_trace(go.Bar(
                    y=df_top_chaco['articulo_label'],
                    x=df_top_chaco['Presupuesto_M'],
                    orientation='h',
                    text=df_top_chaco['Texto'],
                    textposition='outside',
                    cliponaxis=False,
                    marker=dict(color='#e67e22', line=dict(width=0)),
                    hovertemplate='%{customdata}<extra></extra>',
                    customdata=hover_text_pres
                ))
                
                max_pres = df_top_chaco['Presupuesto_M'].max()
                
                fig_presupuesto.update_layout(
                    height=max(400, 20 * 25),
                    margin=dict(t=20, b=25, l=10, r=80),
                    xaxis=dict(visible=False, range=[0, max_pres * 1.2]),
                    yaxis=dict(visible=True, tickfont=dict(size=10)),
                    showlegend=False,
                    plot_bgcolor='white',
                    paper_bgcolor='white'
                )
                
                st.plotly_chart(fig_presupuesto, width='stretch')
                
            with col_g4:
                st.markdown("##### ⏱️ Días de Cobertura (Cap: 7 días)")
                
                # Calcular días de cobertura limitados a 7
                df_top_chaco['dias_cobertura_7'] = df_top_chaco['dias_cobertura'].apply(lambda x: min(x, 7))
                
                # Texto: mostrar STOCK en vez de días
                # df_top_chaco['texto_stock'] = df_top_chaco['STK_CHACO'].apply(lambda x: f"{int(x):,}")
                df_top_chaco['texto_stock'] = df_top_chaco['STK_CHACO'].apply(
                    lambda x: f"{int(x):,}".replace(",", ".") + " uds/stk")

                # Colores según días de cobertura
                df_top_chaco['color_cobertura'] = df_top_chaco['dias_cobertura'].apply(
                    lambda dias: '#27ae60' if dias >= 7 else '#f39c12' if dias >= 3 else '#e67e22' if dias >= 1 else '#e74c3c'
                )
                
                hover_text_cob2 = []
                for idx, row in df_top_chaco.iterrows():
                    texto = f"<b>{row['articulo_label']}</b><br>"
                    texto += f"Cobertura: {row['dias_cobertura']:.1f} días<br>"
                    texto += f"Stock CHACO: {int(row['STK_CHACO']):,} uds<br>"
                    texto += f"Demanda Chaco 1sem: {int(row['chaco_cantidad']):,}"
                    hover_text_cob2.append(texto)
                
                fig_cob2 = go.Figure()
                
                fig_cob2.add_trace(go.Bar(
                    y=df_top_chaco['articulo_label'],
                    x=df_top_chaco['dias_cobertura_7'],
                    orientation='h',
                    text=df_top_chaco['texto_stock'],
                    textposition='outside',
                    cliponaxis=False,
                    marker=dict(color=df_top_chaco['color_cobertura'], line=dict(width=0)),
                    hovertemplate='%{customdata}<extra></extra>',
                    customdata=hover_text_cob2
                ))
                
                # Líneas verticales DIARIAS (1d, 2d, 3d, 4d, 5d, 6d, 7d)
                for dia in range(1, 8):
                    color_linea = '#27ae60' if dia == 7 else '#e67e22'
                    fig_cob2.add_vline(
                        x=dia, line_dash="dash", line_color=color_linea, line_width=1.5, opacity=0.6,
                        annotation_text=f"{dia}d", annotation_position="top",
                        annotation_font_size=9, annotation_font_color=color_linea
                    )
                
                fig_cob2.update_layout(
                    height=max(420, 21 * 25),
                    margin=dict(t=15, b=0, l=20, r=0),
                    xaxis=dict(
                        visible=True, range=[0, 8],
                        title="Días de cobertura",
                        tickmode='array',
                        tickvals=[0, 1, 2, 3, 4, 5, 6, 7],
                        ticktext=['0', '1', '2', '3', '4', '5', '6', '7'],
                        tickfont=dict(size=9)
                    ),
                    yaxis=dict(visible=True, tickfont=dict(size=10)),
                    showlegend=False,
                    plot_bgcolor='white',
                    paper_bgcolor='white'
                )
                
                st.plotly_chart(fig_cob2, width='stretch')
              
            # ═══════════════════════════════════════════════════════════════════
            # GRÁFICOS 3: PRESUPUESTO CORRIENTES 1 SEMANA + COBERTURA
            # ═══════════════════════════════════════════════════════════════════
            
            st.markdown("---")
            st.subheader("🟦 Análisis de Presupuesto CORRIENTES - 1 Semana")
            
            df_top_corr = df_final.nlargest(20, 'presupuesto_corrientes_1sem').copy()
            df_top_corr = df_top_corr.iloc[::-1]
            
            df_top_corr['articulo_label'] = df_top_corr.apply(
                lambda row: f"{row['idartalfa']} - {row['descripcion'][:30]}..." 
                if len(str(row['descripcion'])) > 30 
                else f"{row['idartalfa']} - {row['descripcion']}", 
                axis=1
            )
            
            df_top_corr['dias_cobertura'] = df_top_corr.apply(
                lambda row: (row['stk_corrientes'] / (row['corr_cantidad'] / 7)) 
                if row['corr_cantidad'] > 0 else 0,
                axis=1
            )
            df_top_corr['cobertura_visual'] = df_top_corr['dias_cobertura'].apply(lambda x: min(x, 31))
            df_top_corr['cobertura_texto'] = df_top_corr['dias_cobertura'].apply(
                lambda x: f"{x:.0f}d" if x <= 31 else f"31d+"
            )
            df_top_corr['color_cobertura'] = df_top_corr['dias_cobertura'].apply(get_color_cobertura)
            
            col_g5, col_g6 = st.columns(2)
            
            with col_g5:
                st.markdown("##### 💵 Top 20 por Presupuesto CORRIENTES y unidades a reabastecer")
                
                # df_top_corr['Presupuesto_M'] = df_top_corr['presupuesto_corrientes_1sem'] / 1_000_000
                # df_top_corr['Texto'] = df_top_corr['presupuesto_corrientes_1sem'].apply(
                #     lambda x: f"${x:,.0f}".replace(",", ".")
                # )
                df_top_corr['Presupuesto_M'] = df_top_corr['presupuesto_corrientes_1sem'] / 1_000_000

                # Construir texto con presupuesto y unidades a reabastecer
                df_top_corr['Texto'] = df_top_corr.apply(
                    lambda row: f"${row['presupuesto_corrientes_1sem']:,.0f}".replace(",", ".") + 
                                f" | {int(row['corr_abastecer_1sem']):,} u".replace(",", "."),
                    axis=1
                )

                hover_text_corr = []
                for idx, row in df_top_corr.iterrows():
                    texto = f"<b>{row['articulo_label']}</b><br>"
                    texto += f"Presupuesto Corrientes: ${row['presupuesto_corrientes_1sem']:,.2f}<br>"
                    texto += f"A reabastecer: {int(row['corr_abastecer_1sem']):,} uds<br>"
                    texto += f"Stock Corrientes actual: {int(row['stk_corrientes']):,}"
                    hover_text_corr.append(texto)
                
                fig_pres_corr = go.Figure()
                
                fig_pres_corr.add_trace(go.Bar(
                    y=df_top_corr['articulo_label'],
                    x=df_top_corr['Presupuesto_M'],
                    orientation='h',
                    text=df_top_corr['Texto'],
                    textposition='outside',
                    cliponaxis=False,
                    marker=dict(color='#3498db', line=dict(width=0)),
                    hovertemplate='%{customdata}<extra></extra>',
                    customdata=hover_text_corr
                ))
                
                max_pres_corr = df_top_corr['Presupuesto_M'].max()
                
                fig_pres_corr.update_layout(
                    height=max(420, 21 * 25),
                    margin=dict(t=20, b=25, l=10, r=80),
                    xaxis=dict(visible=False, range=[0, max_pres_corr * 1.2]),
                    yaxis=dict(visible=True, tickfont=dict(size=10)),
                    showlegend=False,
                    plot_bgcolor='white',
                    paper_bgcolor='white'
                )
                
                st.plotly_chart(fig_pres_corr, width='stretch')

            with col_g6:
                st.markdown("##### ⏱️ Días de Cobertura (Cap: 7 días)")
                
                # Calcular días de cobertura limitados a 7
                df_top_corr['dias_cobertura_7'] = df_top_corr['dias_cobertura'].apply(lambda x: min(x, 7))
                
                # Texto: mostrar STOCK en vez de días
                # df_top_corr['texto_stock'] = df_top_corr['stk_corrientes'].apply(lambda x: f"{int(x):,}")
                df_top_corr['texto_stock'] = df_top_corr['stk_corrientes'].apply(
                    lambda x: f"{int(x):,}".replace(",", ".") + " uds/stk")

                # Colores según días de cobertura
                df_top_corr['color_cobertura'] = df_top_corr['dias_cobertura'].apply(
                    lambda dias: '#27ae60' if dias >= 7 else '#f39c12' if dias >= 3 else '#e67e22' if dias >= 1 else '#e74c3c'
                )
                
                hover_text_cob3 = []
                for idx, row in df_top_corr.iterrows():
                    texto = f"<b>{row['articulo_label']}</b><br>"
                    texto += f"Cobertura: {row['dias_cobertura']:.1f} días<br>"
                    texto += f"Stock Corrientes: {int(row['stk_corrientes']):,} uds<br>"
                    texto += f"Demanda Corrientes 1sem: {int(row['corr_cantidad']):,}"
                    hover_text_cob3.append(texto)
                
                fig_cob3 = go.Figure()
                
                fig_cob3.add_trace(go.Bar(
                    y=df_top_corr['articulo_label'],
                    x=df_top_corr['dias_cobertura_7'],
                    orientation='h',
                    text=df_top_corr['texto_stock'],
                    textposition='outside',
                    cliponaxis=False,
                    marker=dict(color=df_top_corr['color_cobertura'], line=dict(width=0)),
                    hovertemplate='%{customdata}<extra></extra>',
                    customdata=hover_text_cob3
                ))
                
                # Líneas verticales DIARIAS (1d, 2d, 3d, 4d, 5d, 6d, 7d)
                for dia in range(1, 8):
                    color_linea = '#27ae60' if dia == 7 else '#e67e22'
                    fig_cob3.add_vline(
                        x=dia, line_dash="dash", line_color=color_linea, line_width=1.5, opacity=0.6,
                        annotation_text=f"{dia}d", annotation_position="top",
                        annotation_font_size=9, annotation_font_color=color_linea
                    )
                
                fig_cob3.update_layout(
                    height=max(420, 21 * 25),
                    margin=dict(t=15, b=0, l=20, r=0),
                    xaxis=dict(
                        visible=True, range=[0, 8],
                        title="Días de cobertura",
                        tickmode='array',
                        tickvals=[0, 1, 2, 3, 4, 5, 6, 7],
                        ticktext=['0', '1', '2', '3', '4', '5', '6', '7'],
                        tickfont=dict(size=9)
                    ),
                    yaxis=dict(visible=True, tickfont=dict(size=10)),
                    showlegend=False,
                    plot_bgcolor='white',
                    paper_bgcolor='white'
                )
                
                st.plotly_chart(fig_cob3, width='stretch')

            
            # ═══════════════════════════════════════════════════════════════════
            # GENERAR Y DESCARGAR EXCEL
            # ═══════════════════════════════════════════════════════════════════
            
            st.markdown("---")
            
            fecha_str = datetime.now().strftime('%d%b%Y')
            nombre_archivo = f"{proveedor_seleccionado.replace(' ', '_')}_presupuesto_1a4sem_{fecha_str}.xlsx"
            
            excel_bytes = crear_libro_excel_formateado(
                df_final, 
                nombres_bloques,
                proveedor_seleccionado
            )
            
            st.download_button(
                label="📥 Descargar Reporte Excel Completo (1-4 semanas)",
                data=excel_bytes,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch'
            )
            
            st.success("✅ Análisis completado exitosamente!")